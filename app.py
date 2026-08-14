# app.py — Détection sons vidéo + mapping interne + SACEM

import os
import re
import time
import hmac
import base64
import hashlib
import tempfile
import subprocess
import unicodedata
import json
from pathlib import Path
from io import BytesIO
from typing import Dict, List, Optional

import pandas as pd
import requests
import streamlit as st
from dotenv import load_dotenv
from rapidfuzz import fuzz, process

from sacem_agent import SacemAgent
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from conductor_order import apply_conductor_order
from ai_realign import ai_realign_with_conductor

load_dotenv(dotenv_path=Path(__file__).with_name(".env"), override=True)
load_dotenv()
ACR_HOST = (os.getenv("ACR_HOST") or "").strip()
ACR_ACCESS_KEY = (os.getenv("ACR_ACCESS_KEY") or "").strip()
ACR_ACCESS_SECRET = (os.getenv("ACR_ACCESS_SECRET") or "").strip()
APP_PASSWORD = (os.getenv("APP_PASSWORD") or "").strip()

st.set_page_config(page_title="Détection sons vidéo", layout="wide")
st.title("Détection sons vidéo — ACRCloud + Mapping interne + SACEM")


# ============================================================
# AUTH
# ============================================================
if APP_PASSWORD:
    st.sidebar.markdown("## Connexion")
    pwd = st.sidebar.text_input("Mot de passe", type="password")
    if pwd != APP_PASSWORD:
        st.warning("Accès protégé.")
        st.stop()


# ============================================================
# UTILS
# ============================================================
def log(msg: str) -> None:
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    st.session_state.setdefault("debug_logs", []).append(line)


def debug_panel() -> None:
    with st.expander("Debug"):
        logs = st.session_state.get("debug_logs", [])
        st.text("\n".join(logs[-300:]))

def to_m6_excel_bytes(df: pd.DataFrame) -> bytes:
    bio = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Relevé"

    yellow = PatternFill("solid", fgColor="FFC000")
    blue = PatternFill("solid", fgColor="BDD7EE")
    red = PatternFill("solid", fgColor="FF0000")
    white_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:J1")
    ws["A1"] = "RELEVE DE DROITS D'AUTEUR"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = "(Merci de nous retourner obligatoirement ce document)"
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["B5"] = "Réalisateur :"
    ws["B5"].font = Font(color="FF0000", bold=True)
    ws["B6"] = "Chaîne : M6"
    ws["B7"] = "Code product (si connu) :"

    ws["I5"] = "Titre du programme :"
    ws["I5"].font = bold
    ws["I6"] = "N° déclinaison :"
    ws["I7"] = "Date de diffusion :"
    ws["I8"] = "Informations complémentaires :"
    ws["I8"].fill = yellow

    headers = [
        "Lien associé à la ligne (colonne B)",
        "TITRE",
        "N° de produit (si connu)",
        "GENRE SACEM",
        "DUREE\n(HH:MM:SS:ii)",
        "TC IN",
        "TC OUT",
        "AUTEUR(S)",
        "COMPOSITEUR(S)",
        "EDITEUR(S)",
        "SOUS-EDITEUR(S)",
        "ISWC",
    ]

    start_row = 10

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col_idx)
        cell.value = header
        cell.fill = yellow
        cell.font = bold
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        values = [
            "",
            row.get("TITRE", ""),
            "",
            row.get("GENRE SACEM", "FDS-Fond sonore"),
            row.get("DUREE", ""),
            row.get("TIME IN", row.get("TC IN", "")),
            row.get("TIME OUT", row.get("TC OUT", "")),
            row.get("AUTEUR(S) SACEM", row.get("AUTEUR(S)", "")),
            row.get("COMPOSITEUR(S) SACEM", row.get("COMPOSITEUR(S)", "")),
            row.get("EDITEUR(S) SACEM", row.get("EDITEUR(S)", "")),
            row.get("SOUS-EDITEUR(S) SACEM", row.get("SOUS-EDITEUR(S)", "")),
            row.get("ISWC", row.get("iswc", "")),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(r_idx, col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if col_idx in [2, 4, 5, 6, 7]:
                cell.fill = blue

    widths = {
        "A": 32,
        "B": 70,
        "C": 16,
        "D": 18,
        "E": 18,
        "F": 16,
        "G": 16,
        "H": 35,
        "I": 35,
        "J": 35,
        "K": 35,
        "L": 20,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[10].height = 38

    wb.save(bio)
    return bio.getvalue()

def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Detection") -> bytes:
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return bio.getvalue()


def norm_spaces(s: str) -> str:
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("\u00A0", " ")
    return re.sub(r"\s+", " ", s).strip()


def normalize_title_decl(s: str) -> str:
    s = norm_spaces(s)
    if not s:
        return ""

    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"\[[^\]]*\]", " ", s)
    s = re.sub(r"\{[^}]*\}", " ", s)
    s = re.sub(r"\.(mp3|wav|aif|aiff|m4a|mp4|mov|mkv|m4v)$", "", s, flags=re.I)
    s = re.split(r"\s[—–-]\s", s, maxsplit=1)[0]
    s = s.strip(" -—–_\t")
    return norm_spaces(s).upper()


def normalize_match_key(s: str) -> str:
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.upper()
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"\[[^\]]*\]", " ", s)
    s = re.sub(r"\{[^}]*\}", " ", s)
    s = re.sub(r"\.(MP3|WAV|AIF|AIFF|M4A|MP4|MOV|MKV|M4V)$", " ", s)
    s = re.sub(
        r"\b(MASTER|V\d+|KEY|OFFICIAL|VIDEO|AUDIO|INSTRUMENTAL|REMIX|REMASTER|VERSION)\b",
        " ",
        s,
    )
    s = re.sub(r"^\d+\s*[-_]", " ", s)
    s = re.sub(r"[_\-\/]+", " ", s)
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def sec_to_timecode(sec: int) -> str:
    sec = max(0, int(sec))
    h = sec // 3600
    m = (sec % 3600) // 60
    s = sec % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def tc_to_seconds(tc: str) -> int:
    try:
        parts = str(tc).split(":")
        if len(parts) >= 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
    except Exception:
        pass
    return 0

def apply_timing_and_shift_following(df, changed_idx, new_time_in, new_time_out):
    out = df.copy()

    start = tc_to_seconds(new_time_in)
    end = tc_to_seconds(new_time_out)

    if end <= start:
        end = start + 1

    out.at[changed_idx, "start_sec"] = start
    out.at[changed_idx, "end_sec"] = end
    out.at[changed_idx, "TIME IN"] = sec_to_timecode(start)
    out.at[changed_idx, "TIME OUT"] = sec_to_timecode(end)
    out.at[changed_idx, "DUREE"] = sec_to_timecode(end - start)

    idxs = list(out.index)
    pos = idxs.index(changed_idx)
    cursor = end

    for next_idx in idxs[pos + 1:]:
        old_start = int(out.at[next_idx, "start_sec"])
        old_end = int(out.at[next_idx, "end_sec"])
        duration = max(1, old_end - old_start)

        out.at[next_idx, "start_sec"] = cursor
        out.at[next_idx, "end_sec"] = cursor + duration
        out.at[next_idx, "TIME IN"] = sec_to_timecode(cursor)
        out.at[next_idx, "TIME OUT"] = sec_to_timecode(cursor + duration)
        out.at[next_idx, "DUREE"] = sec_to_timecode(duration)

        cursor += duration

    return out

def recompute_following_timings(df: pd.DataFrame, changed_idx, new_time_in: str, new_time_out: str) -> pd.DataFrame:
    out = df.copy()

    if changed_idx not in out.index:
        return out

    start = tc_to_seconds(new_time_in)
    end = tc_to_seconds(new_time_out)

    if end <= start:
        end = start + 1

    out.at[changed_idx, "start_sec"] = start
    out.at[changed_idx, "end_sec"] = end
    out.at[changed_idx, "TIME IN"] = sec_to_timecode(start)
    out.at[changed_idx, "TIME OUT"] = sec_to_timecode(end)
    out.at[changed_idx, "DUREE"] = sec_to_timecode(end - start)

    idx_list = list(out.index)
    pos = idx_list.index(changed_idx)

    cursor = end

    for idx in idx_list[pos + 1:]:
        old_start = int(out.at[idx, "start_sec"])
        old_end = int(out.at[idx, "end_sec"])
        duration = max(1, old_end - old_start)

        out.at[idx, "start_sec"] = cursor
        out.at[idx, "end_sec"] = cursor + duration
        out.at[idx, "TIME IN"] = sec_to_timecode(cursor)
        out.at[idx, "TIME OUT"] = sec_to_timecode(cursor + duration)
        out.at[idx, "DUREE"] = sec_to_timecode(duration)

        cursor = cursor + duration

    return out

def find_audio_ref_file(acr_title: str, hit: Optional[Dict[str, str]] = None) -> str:
    base_dir = Path(__file__).parent / "audio_refs"
    candidates = []

    if hit:
        candidates.append(hit.get("audio_ref", ""))
        candidates.append(hit.get("CORRESPONDANCE", ""))
        candidates.append(hit.get("TITRE", ""))

    candidates.append(acr_title)

    exts = [".wav", ".mp3", ".aif", ".aiff", ".m4a"]

    for c in candidates:
        c = norm_spaces(c)
        if not c:
            continue

        p = Path(c)

        if p.is_absolute() and p.exists():
            return str(p)

        if base_dir.exists():
            direct = base_dir / p.name
            if direct.exists():
                return str(direct)

            key = normalize_match_key(c)

            for f in base_dir.rglob("*"):
                if f.is_file() and f.suffix.lower() in exts:
                    if normalize_match_key(f.stem) == key:
                        return str(f)

            for f in base_dir.rglob("*"):
                if f.is_file() and f.suffix.lower() in exts:
                    if key and key in normalize_match_key(f.stem):
                        return str(f)

    return ""

def repair_episode_metadata():
    for p in EPISODES_DIR.glob("*.parquet"):
        meta = EPISODES_DIR / f"{p.stem}.json"
        if not meta.exists():
            meta.write_text(json.dumps({
                "episode_id": p.stem,
                "paths": [""],
                "name": p.stem,
                "saved_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            }, ensure_ascii=False, indent=2))

CACHE_DIR = Path("cache")
EPISODES_DIR = CACHE_DIR / "episodes"
repair_episode_metadata()
SACEM_CACHE_DIR = CACHE_DIR / "sacem"
AUDIO_CACHE_DIR = CACHE_DIR / "audio_excerpts"

for d in [EPISODES_DIR, SACEM_CACHE_DIR, AUDIO_CACHE_DIR]:
    d.mkdir(parents=True, exist_ok=True)


def episode_id_from_paths(paths: List[str]) -> str:
    raw = ""
    for p in paths:
        path = Path(p)
        raw += str(path.resolve())
        raw += str(path.stat().st_size) if path.exists() else ""
        raw += str(int(path.stat().st_mtime)) if path.exists() else ""
    return hashlib.md5(raw.encode()).hexdigest()


def episode_files(ep_id: str):
    return {
        "data": EPISODES_DIR / f"{ep_id}.parquet",
        "meta": EPISODES_DIR / f"{ep_id}.json",
    }


def save_episode_cache(ep_id: str, paths: List[str], df: pd.DataFrame):
    files = episode_files(ep_id)
    safe_parquet_df(df).to_parquet(files["data"], index=False)

    meta = {
        "episode_id": ep_id,
        "paths": paths,
        "name": Path(paths[0]).name if paths else ep_id,
        "saved_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }

    files["meta"].write_text(json.dumps(meta, ensure_ascii=False, indent=2))


def load_episode_cache(ep_id: str):
    files = episode_files(ep_id)
    if not files["data"].exists() or not files["meta"].exists():
        return None, None

    df = pd.read_parquet(files["data"])
    meta = json.loads(files["meta"].read_text())
    return df, meta


def list_episode_cache():
    items = []
    for meta_file in EPISODES_DIR.glob("*.json"):
        try:
            meta = json.loads(meta_file.read_text())
            items.append(meta)
        except Exception:
            pass
    return sorted(items, key=lambda x: x.get("saved_at", ""), reverse=True)


def sacem_cache_key(title: str, artist: str) -> str:
    raw = f"{normalize_match_key(title)}||{normalize_match_key(artist)}"
    return hashlib.md5(raw.encode()).hexdigest()


def cached_sacem_lookup(title: str, artist: str) -> dict:
    key = sacem_cache_key(title, artist)
    p = SACEM_CACHE_DIR / f"{key}.json"

    if p.exists():
        return json.loads(p.read_text())

    agent = SacemAgent(headless=True)
    result = agent.search(title, artist)

    p.write_text(json.dumps(result, ensure_ascii=False, indent=2))
    return result

# ============================================================
# ACRCloud
# ============================================================
_SESSION = requests.Session()


def acr_sign(string_to_sign: bytes, secret: str) -> str:
    return base64.b64encode(
        hmac.new(secret.encode(), string_to_sign, hashlib.sha1).digest()
    ).decode()


def recognize_chunk(
    wav_path: str,
    host: str,
    access_key: str,
    access_secret: str,
    timeout=(30, 300),
    retries: int = 4,
):
    wav_file = Path(wav_path)
    if not wav_file.exists():
        raise FileNotFoundError(wav_path)

    http_uri = "/v1/identify"
    timestamp = str(int(time.time()))
    string_to_sign = f"POST\n{http_uri}\n{access_key}\naudio\n1\n{timestamp}"
    sign = acr_sign(string_to_sign.encode(), access_secret)

    data = {
        "access_key": access_key,
        "sample_bytes": wav_file.stat().st_size,
        "timestamp": timestamp,
        "signature": sign,
        "data_type": "audio",
        "signature_version": "1",
    }

    last_exc = None

    for attempt in range(1, retries + 1):
        try:
            with open(wav_path, "rb") as f:
                files = {"sample": f}
                r = _SESSION.post(
                    f"https://{host}{http_uri}",
                    data=data,
                    files=files,
                    timeout=timeout,
                )
            r.raise_for_status()
            return r.json()
        except Exception as e:
            last_exc = e
            time.sleep(1.4 ** (attempt - 1))

    raise RuntimeError(f"ACRCloud failed after retries: {last_exc}")


# ============================================================
# FFMPEG
# ============================================================
def ffprobe_duration_seconds(path: str) -> float:
    cmd = [
        "ffprobe", "-v", "error",
        "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1",
        path,
    ]
    out = subprocess.check_output(cmd).decode("utf-8").strip()
    return float(out) if out else 0.0


def ffmpeg_extract_audio(video_path: str, out_wav_path: str) -> None:
    cmd = [
        "ffmpeg", "-y",
        "-i", video_path,
        "-vn",
        "-ac", "1",
        "-ar", "44100",
        out_wav_path,
    ]
    subprocess.check_call(cmd)


def chunk_audio_overlap(wav_path: str, out_dir: str, seconds: int = 8, step: int = 2) -> None:
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    duration = int(ffprobe_duration_seconds(wav_path))

    idx = 0
    start = 0

    while start < duration:
        out_path = Path(out_dir) / f"chunk_{idx:04d}_{start:06d}.wav"
        cmd = [
            "ffmpeg", "-y",
            "-ss", str(start),
            "-i", wav_path,
            "-t", str(seconds),
            "-ac", "1",
            "-ar", "44100",
            str(out_path),
        ]
        subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        idx += 1
        start += int(step)


def compute_part_offsets(paths: List[str]) -> Dict[str, int]:
    offsets = {}
    cum_sec = 0

    for i, path in enumerate(paths, start=1):
        part = f"P{i}"
        offsets[part] = cum_sec
        try:
            cum_sec += int(ffprobe_duration_seconds(path))
        except Exception:
            pass

    return offsets


def prepare_audio_excerpts(
    df_occ: pd.DataFrame,
    video_paths: Dict[str, str],
    excerpt_dir: str = "audio_excerpts",
) -> pd.DataFrame:
    Path(excerpt_dir).mkdir(exist_ok=True)
    out = df_occ.copy()
    audio_paths = []

    for idx, row in out.iterrows():
        part = str(row.get("part", "P1"))
        video_path = video_paths.get(part)

        start_sec = int(row.get("start_sec", 0))
        end_sec = int(row.get("end_sec", start_sec + 8))
        duration = max(3, end_sec - start_sec)

        safe_title = normalize_match_key(row.get("TITRE", f"item_{idx}"))[:40]
        audio_path = Path(excerpt_dir) / f"{part}_{idx}_{start_sec}_{safe_title}.wav"

        if video_path and not audio_path.exists():
            cmd = [
                "ffmpeg", "-y",
                "-ss", str(start_sec),
                "-i", video_path,
                "-t", str(duration),
                "-vn",
                "-ac", "1",
                "-ar", "44100",
                str(audio_path),
            ]
            subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

        audio_paths.append(str(audio_path) if audio_path.exists() else "")

    out["audio_excerpt"] = audio_paths
    return out


# ============================================================
# MAPPING
# ============================================================
def load_mapping(uploaded_file) -> pd.DataFrame:
    if uploaded_file.name.lower().endswith(".csv"):
        return pd.read_csv(uploaded_file)

    return pd.read_excel(uploaded_file)


def video_hash(video_path: str) -> str:
    p = Path(video_path)

    txt = (
        str(p.resolve())
        + str(p.stat().st_size)
        + str(int(p.stat().st_mtime))
    )

    return hashlib.md5(txt.encode()).hexdigest()

def build_mapping_index(mapping_df: pd.DataFrame) -> Dict[str, Dict[str, str]]:
    idx: Dict[str, Dict[str, str]] = {}

    if mapping_df is None or mapping_df.empty or len(mapping_df.columns) < 2:
        return idx

    col_final = mapping_df.columns[0]
    col_alias = mapping_df.columns[1]
    col_audio = mapping_df.columns[2] if len(mapping_df.columns) >= 3 else None

    for _, r in mapping_df.iterrows():
        final_title = norm_spaces(r.get(col_final, ""))
        aliases_raw = norm_spaces(r.get(col_alias, ""))

        if not final_title or final_title.lower() == "nan":
            continue

        aliases = [a.strip() for a in aliases_raw.split("/") if a.strip()]
        aliases.append(final_title)

        for alias in aliases:
            key = normalize_match_key(alias)
            if not key:
                continue

            audio_ref = ""
            if col_audio is not None:
                audio_ref = Path(str(r.get(col_audio, "")).strip()).name

            idx[key] = {
                "TITRE": normalize_title_decl(final_title),
                "CORRESPONDANCE": alias,
                "audio_ref": audio_ref,
            }

    return idx


def fuzzy_match_mapping(acr_title: str, mapping_index: Dict[str, Dict[str, str]], threshold: int = 82):
    if not acr_title or not mapping_index:
        return None, 0, ""

    query = normalize_match_key(acr_title)
    choices = list(mapping_index.keys())

    best = process.extractOne(query, choices, scorer=fuzz.partial_ratio)

    if not best:
        return None, 0, ""

    best_key, score, _ = best

    if score >= threshold:
        return mapping_index.get(best_key), int(score), best_key

    return None, int(score), best_key


# ============================================================
# AGGREGATION
# ============================================================
def aggregate_video_occurrences(
    df_hits: pd.DataFrame,
    chunk_seconds: int,
    gap_tolerance_sec: int = 4,
) -> pd.DataFrame:
    if df_hits.empty:
        return df_hits

    df = df_hits.copy()
    df["start_sec"] = pd.to_numeric(df["start_sec"], errors="coerce").fillna(0).astype(int)
    df["end_sec"] = df["start_sec"] + int(chunk_seconds)

    df["match_title"] = df["final_title"].fillna(df["titre"]).astype(str)
    df["key"] = df["part"].astype(str) + "||" + df["match_title"].str.upper().str.strip()

    df = df.sort_values(["part", "match_title", "start_sec"]).reset_index(drop=True)

    segment_rows = []

    for _, g in df.groupby("key", sort=False):
        g = g.sort_values("start_sec").reset_index(drop=True)

        cur_start = int(g.loc[0, "start_sec"])
        cur_end = int(g.loc[0, "end_sec"])
        first = g.loc[0].to_dict()

        scores = []
        if "score" in g.columns:
            scores = pd.to_numeric(g["score"], errors="coerce").dropna().tolist()

        acr_titles = []
        for x in g["titre"].astype(str).tolist():
            if x and x not in acr_titles:
                acr_titles.append(x)

        def add_segment(start, end):
            segment_rows.append({
                "part": first.get("part", ""),
                "TITRE": first.get("final_title") or first.get("titre", ""),
                "ARTISTE": first.get("artiste", ""),
                "SOURCE TITLE": " | ".join(acr_titles),
                "TIME IN": sec_to_timecode(start),
                "TIME OUT": sec_to_timecode(end),
                "DUREE": sec_to_timecode(end - start),
                "MATCH SCORE": first.get("mapping_score", 0),
                "ACR SCORE": max(scores) if scores else "",
                "start_sec": start,
                "end_sec": end,
                "audio_ref": first.get("audio_ref", ""),
                "isrc": first.get("isrc", ""),
            })

        for i in range(1, len(g)):
            s = int(g.loc[i, "start_sec"])
            e = int(g.loc[i, "end_sec"])

            if s <= cur_end + int(gap_tolerance_sec):
                cur_end = max(cur_end, e)
            else:
                add_segment(cur_start, cur_end)
                cur_start = s
                cur_end = e

        add_segment(cur_start, cur_end)

    out = pd.DataFrame(segment_rows)
    out = out.sort_values(["part", "start_sec"]).reset_index(drop=True)
    return out


# ============================================================
# SACEM
# ============================================================
@st.cache_data(show_spinner=False)
def cached_sacem_search(title: str, artist: str) -> dict:
    agent = SacemAgent(headless=True)
    return agent.search(title, artist)


def enrich_df_with_sacem(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    for col in [
        "SACEM STATUT",
        "AUTEUR(S) SACEM",
        "COMPOSITEUR(S) SACEM",
        "EDITEUR(S) SACEM",
        "SOUS-EDITEUR(S) SACEM",
        "INTERPRETE(S) SACEM",
        "URL SACEM",
        "URL RECHERCHE SACEM",
    ]:
        if col not in out.columns:
            out[col] = "" 

    total = len(out)
    progress = st.progress(0.0)
    status_box = st.empty()

    for pos, idx in enumerate(out.index, start=1):

        title = str(out.at[idx, "TITRE"]).strip()

        artist = ""
        if "ARTISTE" in out.columns:
            artist = str(out.at[idx, "ARTISTE"]).strip()

        log(f"SACEM SEARCH title='{title}' artist='{artist}'")

        status_box.write(f"SACEM {pos}/{total} : **{title}**")

        if not title:
            progress.progress(pos / total)
            continue    

        try:
          res = cached_sacem_lookup(title, artist)
          
          out.at[idx, "SACEM STATUT"] = res.get("status", "")
          out.at[idx, "AUTEUR(S) SACEM"] = "; ".join(res.get("authors", []))
          out.at[idx, "COMPOSITEUR(S) SACEM"] = "; ".join(res.get("composers", []))
          out.at[idx, "EDITEUR(S) SACEM"] = "; ".join(res.get("publishers", []))
          out.at[idx, "SOUS-EDITEUR(S) SACEM"] = "; ".join(res.get("sub_publishers", []))
          out.at[idx, "INTERPRETE(S) SACEM"] = "; ".join(res.get("performers", []))
          out.at[idx, "SACEM TITLE"] = res.get("title", "")
          out.at[idx, "SACEM SCORE"] = res.get("title_match_score", "")
          out.at[idx, "URL RECHERCHE SACEM"] = res.get("search_url", "")
          url = str(res.get("url", "")).strip()

          if url.startswith("https://repertoire.sacem.fr/detail-oeuvre/"):
             out.at[idx, "URL SACEM"] = url
          else:
             out.at[idx, "URL SACEM"] = ""
             log(f"SACEM URL title='{title}' url='{res.get('url', '')}'")
             log(
    f"SACEM RESULT input='{title}' "
    f"sacem_title='{res.get('title', '')}' "
    f"score='{res.get('title_match_score', '')}' "
    f"url='{res.get('url', '')}'"
)

        except Exception as e:
            out.at[idx, "SACEM STATUT"] = "error"
            out.at[idx, "URL SACEM"] = str(e)
            log(f"SACEM error {title}: {repr(e)}")

        progress.progress(pos / total)

    status_box.empty()
    return out


# ============================================================
# SAFE LISTING
# ============================================================
def safe_list_dirs(root: Path) -> List[Path]:
    out: List[Path] = []

    try:
        for p in root.iterdir():
            try:
                if not p.is_dir():
                    continue
                if p.name.startswith(".") or p.name.startswith("#"):
                    continue
                if p.name.lower() in {"#recycle", "@eadir"}:
                    continue
                out.append(p)
            except Exception:
                continue
    except Exception:
        return []

    return sorted(out, key=lambda x: x.name.lower())

def safe_parquet_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    text_cols = [
        "ACR SCORE",
        "MATCH SCORE",
        "isrc",
        "audio_ref",
        "audio_excerpt",
        "SACEM STATUT",
        "SACEM TITLE",
        "SACEM SCORE",
        "AUTEUR(S) SACEM",
        "COMPOSITEUR(S) SACEM",
        "EDITEUR(S) SACEM",
        "SOUS-EDITEUR(S) SACEM",
        "INTERPRETE(S) SACEM",
        "URL SACEM",
        "URL RECHERCHE SACEM",
        "ISWC",
        "DURÉE HIST MIN",
        "DURÉE HIST MAX",
        "DURÉE HIST AVG",
        "DURÉE HIST MEDIAN",
        "ACR TITLE ORIGINAL",
        "CONDUCTEUR TITLE",
        "CONDUCTEUR WARNING",
    ]
    for col in text_cols:
        if col in out.columns:
            out[col] = out[col].fillna("").astype(str)

    for col in ["start_sec", "end_sec"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0).astype(int)

    return out

def safe_rglob_videos(root: Path) -> List[Path]:
    files = []

    for ext in ["*.mp4", "*.mov", "*.mkv", "*.m4v"]:
        try:
            files.extend(root.rglob(ext))
        except Exception:
            pass

    return sorted(files)

def make_audio_excerpt(video_path: str, start_sec: int, duration_sec: int, out_dir: str) -> str:
    Path(out_dir).mkdir(parents=True, exist_ok=True)

    safe_start = max(0, int(start_sec))
    safe_duration = max(1, int(duration_sec))

    key = hashlib.md5(f"{video_path}_{safe_start}_{safe_duration}".encode()).hexdigest()
    out_path = Path(out_dir) / f"manual_excerpt_{key}.wav"

    cmd = [
        "ffmpeg", "-y",
        "-ss", str(safe_start),
        "-i", video_path,
        "-t", str(safe_duration),
        "-vn",
        "-ac", "1",
        "-ar", "44100",
        str(out_path),
    ]

    subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    return str(out_path)
def apply_timing_and_shift_following(df, changed_idx, new_time_in, new_time_out):
    out = df.copy()

    start = tc_to_seconds(new_time_in)
    end = tc_to_seconds(new_time_out)

    if end <= start:
        end = start + 1

    out.at[changed_idx, "start_sec"] = start
    out.at[changed_idx, "end_sec"] = end
    out.at[changed_idx, "TIME IN"] = sec_to_timecode(start)
    out.at[changed_idx, "TIME OUT"] = sec_to_timecode(end)
    out.at[changed_idx, "DUREE"] = sec_to_timecode(end - start)

    idxs = list(out.index)
    pos = idxs.index(changed_idx)
    cursor = end

    for next_idx in idxs[pos + 1:]:
        old_start = int(out.at[next_idx, "start_sec"])
        old_end = int(out.at[next_idx, "end_sec"])
        duration = max(1, old_end - old_start)

        out.at[next_idx, "start_sec"] = cursor
        out.at[next_idx, "end_sec"] = cursor + duration
        out.at[next_idx, "TIME IN"] = sec_to_timecode(cursor)
        out.at[next_idx, "TIME OUT"] = sec_to_timecode(cursor + duration)
        out.at[next_idx, "DUREE"] = sec_to_timecode(duration)

        cursor += duration

    return out

# ============================================================
# SIDEBAR
# ============================================================
with st.sidebar:
    st.header("Paramètres")

    debug_enabled = st.checkbox("Activer debug", value=True)

    st.markdown("---")
    st.markdown("## Mapping interne")
    st.caption("Colonne A = titre final, colonne B = intitulés/aliases séparés par /, colonne C = audio ref optionnel")
    mapping_file = st.file_uploader("Uploader mapping Excel/CSV", type=["xlsx", "csv"])

    st.markdown("---")
    fuzzy_threshold = st.slider("Seuil fuzzy match", 50, 100, 82, 1)

mapping_df = None
mapping_index: Dict[str, Dict[str, str]] = {}

if mapping_file is not None:
    try:
        mapping_df = load_mapping(mapping_file)
        mapping_index = build_mapping_index(mapping_df)
        st.sidebar.success(f"Mapping chargé : {len(mapping_index)} alias")
    except Exception as e:
        st.sidebar.error(f"Erreur mapping : {e}")
        mapping_index = {}

if not (ACR_HOST and ACR_ACCESS_KEY and ACR_ACCESS_SECRET):
    st.error("ACRCloud non configuré. Mets ACR_HOST, ACR_ACCESS_KEY, ACR_ACCESS_SECRET dans .env")
    st.stop()


# ============================================================
# CONTROLS
# ============================================================
c1, c2 = st.columns([1, 1])

if c1.button("Reset session"):
    for k in [
        "results_df",
        "video_detected_df",
        "video_selection_df",
        "video_current_path",
        "video_start_sec",
        "debug_logs",
    ]:
        st.session_state.pop(k, None)
    st.rerun()

if c2.button("Stop"):
    st.session_state["stop"] = True


# ============================================================
# VIDEO MODE
# ============================================================
st.markdown("## Épisodes déjà traités")

episodes = list_episode_cache()

if episodes:
    labels = [
        f"{e.get('saved_at', '')} — {e.get('name', e.get('episode_id'))}"
        for e in episodes
    ]

    selected_label = st.selectbox("Charger un épisode déjà traité", labels)
    selected_meta = episodes[labels.index(selected_label)]

    if st.button("Charger cet épisode"):
        df_cached, meta = load_episode_cache(selected_meta["episode_id"])

        if df_cached is not None:
            st.session_state["episode_id"] = selected_meta["episode_id"]

            paths_meta = meta.get("paths", [])

            if paths_meta:
                st.session_state["video_current_path"] = paths_meta[0]
            else:
                st.session_state.pop("video_current_path", None)
                st.warning(
            "Épisode chargé sans chemin vidéo. Les titres sont disponibles mais la vidéo ne peut pas être affichée."
        )

            st.session_state["video_detected_df"] = df_cached
            st.session_state["video_selection_df"] = df_cached.copy()

            st.success("Épisode chargé depuis l'historique.")
            st.rerun()
 

chunk_seconds = st.slider("Durée des chunks", 4, 20, 8, 1)
chunk_step = st.slider("Pas entre chunks", 1, chunk_seconds, 2, 1)
gap_tolerance = st.slider("Tolérance fusion occurrences", 0, 20, 4, 1)

tab1, tab2, tab3 = st.tabs(["Depuis bibliothèque", "Chemin local", "Uploader vidéo"])

paths: List[str] = []
temp_uploaded_paths: List[str] = []

with tab1:
    if Path("/mnt/FM_FR_LIBRARY").exists():
        LIB_ROOT = Path("/mnt/FM_FR_LIBRARY")
    elif Path("/Volumes/FM_FR_LIBRARY").exists():
        LIB_ROOT = Path("/Volumes/FM_FR_LIBRARY")
    else:
        LIB_ROOT = None

    if LIB_ROOT is None:
        st.info("Bibliothèque non détectée. Utilise l’upload vidéo.")
    else:
        lvl1_dirs = safe_list_dirs(LIB_ROOT)

        if lvl1_dirs:
            lvl1 = st.selectbox("Dossier principal", lvl1_dirs, format_func=lambda p: p.name)
            lvl2_dirs = safe_list_dirs(lvl1)

            if lvl2_dirs:
                lvl2 = st.selectbox("Sous-dossier", lvl2_dirs, format_func=lambda p: p.name)
                videos = safe_rglob_videos(lvl2)

                selected = st.multiselect(
                    "Sélectionner vidéo(s)",
                    videos,
                    format_func=lambda p: str(p.relative_to(LIB_ROOT)),
                )
                paths = [str(p) for p in selected]

with tab2:
    st.caption("Recommandé pour les gros fichiers : l’app lit directement le fichier sur ton disque.")

    local_path_txt = st.text_input(
        "Chemin fichier ou dossier vidéo",
        placeholder="/Users/boussaadiam/Movies/mon_episode.mp4 ou /Volumes/FM_FR_LIBRARY/..."
    )

    if local_path_txt:
        local_path = Path(local_path_txt).expanduser()

        if local_path.is_file():
            paths = [str(local_path)]
        elif local_path.is_dir():
            found = []
            for ext in ["*.mp4", "*.mov", "*.mkv", "*.m4v"]:
                found.extend(local_path.rglob(ext))

            found = sorted(found)
            selected_local = st.multiselect(
                "Vidéos trouvées",
                found,
                format_func=lambda p: str(p),
            )
            paths = [str(p) for p in selected_local]
        else:
            st.error("Chemin introuvable.")

with tab3:
    uploaded_videos = st.file_uploader(
        "Uploader une ou plusieurs vidéos",
        type=["mp4", "mov", "mkv", "m4v"],
        accept_multiple_files=True,
    )

    if uploaded_videos:
        for uf in uploaded_videos:
            suffix = Path(uf.name).suffix or ".mp4"
            tmp_path = tempfile.NamedTemporaryFile(delete=False, suffix=suffix).name
            with open(tmp_path, "wb") as f:
                f.write(uf.getbuffer())
            temp_uploaded_paths.append(tmp_path)

        paths = temp_uploaded_paths

st.markdown("## Conducteur MDP")
st.session_state["use_ai_realign"] = st.checkbox(
    "Utiliser l'IA pour réaligner avec le conducteur",
    value=False,
)
use_conductor_order = st.checkbox(
    "Utiliser le conducteur pour corriger l’ordre des titres",
    value=True,
)

conductor_type = st.selectbox(
    "Type de conducteur",
    ["Lundi-Jeudi", "Vendredi"],
)
 

 

if st.button("Analyser les vidéos", type="primary"):
    for k in ["results_df", "video_detected_df", "video_selection_df"]:
        st.session_state.pop(k, None)

    st.session_state["stop"] = False
    st.session_state["video_start_sec"] = 0

    if not paths:
        cache_dir = Path("cache")
        cache_dir.mkdir(exist_ok=True)

        cache_file = cache_dir / f"{video_hash(paths[0])}.parquet"

        if cache_file.exists():

            st.success("Résultats ACRCloud trouvés dans le cache")

            df_occ = pd.read_parquet(cache_file)

            st.session_state["video_detected_df"] = df_occ
            st.session_state["video_selection_df"] = df_occ.copy()
            ep_id = episode_id_from_paths(paths)
            save_episode_cache(ep_id, paths, df_occ)
            st.session_state["episode_id"] = ep_id

            st.rerun()
        st.error("Sélectionne ou upload au moins une vidéo.")
        st.stop()

    st.session_state["video_current_path"] = paths[0]

    results = []
    part_offsets = compute_part_offsets(paths)

    for part_idx, video_path in enumerate(paths, start=1):
        if st.session_state.get("stop"):
            break

        part_label = f"P{part_idx}"
        log(f"{part_label} video={video_path}")

        wav_path = tempfile.NamedTemporaryFile(suffix=f"_{part_label}.wav", delete=False).name

        with st.spinner(f"{part_label} extraction audio..."):
            ffmpeg_extract_audio(video_path, wav_path)

        chunks_dir = Path(tempfile.mkdtemp(prefix=f"chunks_{part_label}_"))

        with st.spinner(f"{part_label} découpage audio overlap..."):
            chunk_audio_overlap(
                wav_path,
                str(chunks_dir),
                seconds=int(chunk_seconds),
                step=int(chunk_step),
            )

        chunk_files = sorted(chunks_dir.glob("chunk_*.wav"))
        progress = st.progress(0.0)
        status_box = st.empty()

        for n, p in enumerate(chunk_files, start=1):
            if st.session_state.get("stop"):
                break

            status_box.write(f"{part_label} ACRCloud {n}/{len(chunk_files)} : {p.name}")

            try:
                parts = p.stem.split("_")
                start_local = int(parts[2]) if len(parts) >= 3 else 0
                start_sec = start_local + part_offsets.get(part_label, 0)
                r = recognize_chunk(
                     str(p),
            host=ACR_HOST,
            access_key=ACR_ACCESS_KEY,
            access_secret=ACR_ACCESS_SECRET,
                 )
                status = r.get("status", {}) or {}
                md = r.get("metadata", {}) or {}
                custom = md.get("custom_files", []) or []
                music = md.get("music", []) or []

                if n <= 10 or custom or music:
                    log(
                        f"DEBUG ACR chunk={p.name} "
                        f"status={status} "
                        f"custom_count={len(custom)} "
                        f"music_count={len(music)} "
                        f"keys={list(md.keys())}"
                        )
                ok = (status.get("code") == 0) or (str(status.get("msg", "")).lower() == "success")
                if custom:
                     log(f"DEBUG custom first={json.dumps(custom[0], ensure_ascii=False)[:1000]}")

                if music:
                  log(f"DEBUG music first={json.dumps(music[0], ensure_ascii=False)[:1000]}")
               
                ok = (status.get("code") == 0) or (str(status.get("msg", "")).lower() == "success")

                if not ok:
                    msg = str(status.get("msg", "")).lower()
                    code = status.get("code")

                    if code == 3003:
                         st.error("Limite ACRCloud atteinte.")
                         break

                    if "no result" in msg or code in [1001, 2004]:
                        results.append({
                            "part": part_label,
                            "titre": "UNKNOWN",
                            "artiste": "",
                            "source": "unknown",
                            "score": "",
                            "isrc": "",
                            "start_sec": int(start_sec),
                            "mapping_found": False,
                            "mapping_score": 0,
                            "mapping_key": "",
                            "final_title": "UNKNOWN",
                            "audio_ref": "",
                        })

                    progress.progress(n / len(chunk_files))
                    continue

                if not custom and not music:
                    log(f"DEBUG skip no metadata match chunk={p.name} status={status} md_keys={list(md.keys())}")
                    progress.progress(n / len(chunk_files))
                    continue
                if not ok:
                    progress.progress(n / len(chunk_files))
                    continue

                md = r.get("metadata", {}) or {}
                custom = md.get("custom_files", []) or []
                music = md.get("music", []) or []

                source = "custom_files" if custom else "music"
                items = custom if custom else music

                if not items:
                    progress.progress(n / len(chunk_files))
                    continue

                m = items[0]
                title = norm_spaces(m.get("title", ""))

                if not title:
                    progress.progress(n / len(chunk_files))
                    continue

                artist = ""
                if source == "music":
                    artists = m.get("artists") or []
                    if artists and isinstance(artists, list):
                        artist = norm_spaces(artists[0].get("name", ""))

                hit, match_score, match_key = fuzzy_match_mapping(
                    title,
                    mapping_index,
                    threshold=int(fuzzy_threshold),
                )

                final_title = hit.get("TITRE", title) if hit else title
                audio_ref = find_audio_ref_file(title, hit)

                results.append({
                    "part": part_label,
                    "titre": title,
                    "artiste": artist,
                    "source": source,
                    "score": m.get("score", ""),
                    "isrc": ((m.get("external_ids") or {}).get("isrc")) or m.get("isrc", ""),
                    "start_sec": int(start_sec),
                    "mapping_found": bool(hit),
                    "mapping_score": match_score,
                    "mapping_key": match_key,
                    "final_title": normalize_title_decl(final_title),
                    "audio_ref": audio_ref,
                })

            except Exception as e:
                log(f"Erreur chunk {p.name}: {repr(e)}")

            progress.progress(n / len(chunk_files))

    df_hits = pd.DataFrame(results)

    if df_hits.empty:
        st.warning("Aucun titre détecté.")
    else:
        df_occ = aggregate_video_occurrences(
            df_hits,
            chunk_seconds=int(chunk_seconds),
            gap_tolerance_sec=int(gap_tolerance),
        )
        if use_conductor_order:
            df_occ = apply_conductor_order(df_occ, conductor_type)

            for idx, row in df_occ.iterrows():
                 title = str(row.get("TITRE", "")).strip()

                 hit, _, _ = fuzzy_match_mapping(
                    title,
                    mapping_index,
                    threshold=70,
                )

                 df_occ.at[idx, "audio_ref"] = find_audio_ref_file(title, hit)
            if st.session_state.get("use_ai_realign", False):
                df_occ = ai_realign_with_conductor(df_occ, conductor_type)

                video_paths_by_part = {f"P{i+1}": p for i, p in enumerate(paths)}

        with st.spinner("Préparation des extraits audio détectés..."):
            video_paths_by_part = {
    f"P{i+1}": p
    for i, p in enumerate(paths)
}
            df_occ = prepare_audio_excerpts(df_occ, video_paths_by_part)

        st.session_state["video_detected_df"] = df_occ
        cache_dir = Path("cache")
        cache_dir.mkdir(exist_ok=True)

        cache_file = cache_dir / f"{video_hash(paths[0])}.parquet"

        safe_parquet_df(df_occ).to_parquet(cache_file)
        st.session_state["video_selection_df"] = df_occ.copy()
        st.success(f"{len(df_occ)} occurrence(s) détectée(s).")


# ============================================================
# VIDEO SOURCE
# ============================================================
video_path = st.session_state.get("video_current_path", "")

if video_path and Path(video_path).exists():
    st.markdown("## Vidéo source")

    st.video(
        video_path,
        start_time=int(st.session_state.get("video_start_sec", 0)),
    )

elif "video_current_path" in st.session_state:
    st.warning("Vidéo source introuvable ou vide. Les résultats sont chargés, mais la vidéo ne peut pas être affichée.")


# ============================================================
# RESULTATS
# ============================================================
if "video_detected_df" in st.session_state:
    st.markdown("## Titres détectés")

    df_sel = st.session_state["video_selection_df"].copy()
    df_sel = df_sel[
    df_sel["TITRE"].astype(str).str.upper().str.strip() != "UNKNOWN"
     ].copy()

    st.session_state["video_selection_df"] = df_sel
    st.session_state["video_detected_df"] = df_sel.copy()

    if "garder" not in df_sel.columns:
        df_sel["garder"] = True

    edited = st.data_editor(
        df_sel,
        key="video_editor",
        width="stretch",
        num_rows="dynamic",
    )

    st.session_state["video_selection_df"] = edited

    c_sacem1, c_sacem2 = st.columns([1, 2])

    if c_sacem1.button("Enrichir via SACEM", type="primary"):
        with st.spinner("Recherche SACEM en cours..."):
            enriched = enrich_df_with_sacem(st.session_state["video_selection_df"])
            st.session_state["video_selection_df"] = enriched
            st.session_state["video_detected_df"] = enriched.copy()
            ep_id = st.session_state.get("episode_id")
            if ep_id:
             paths_to_save = [st.session_state.get("video_current_path", "")]
             save_episode_cache(ep_id, paths_to_save, enriched)
             st.success("Enrichissement SACEM terminé.")
        st.rerun()

    c_sacem2.caption("SACEM est lancé une fois par occurrence agrégée, pas par chunk.")

    st.markdown("## Écoute rapide")

    if edited.empty:
        st.warning("Aucune occurrence à afficher.")
    else:
        for idx, row in edited.iterrows():
            titre = str(row.get("TITRE", "")).strip()
            artiste = str(row.get("ARTISTE", "")).strip()
            time_in = str(row.get("TIME IN", "00:00:00")).strip()
            time_out = str(row.get("TIME OUT", "00:00:00")).strip()
            duree = str(row.get("DUREE", "")).strip()
            source_title = str(row.get("SOURCE TITLE", "")).strip()
            audio_ref = str(row.get("audio_ref", "")).strip()
            isrc = str(row.get("isrc", "")).strip()
        

            with st.expander(f"{time_in} — {titre}", expanded=False):
                edit_cols = st.columns(4)

                new_time_in = edit_cols[0].text_input(
                    "TIME IN",
                    value=time_in,
                    key=f"edit_time_in_{idx}",
                )

                new_time_out = edit_cols[1].text_input(
                    "TIME OUT",
                    value=time_out,
                    key=f"edit_time_out_{idx}",
                )

                if not duree:
                    duree = sec_to_timecode(
                        max(0, tc_to_seconds(time_out) - tc_to_seconds(time_in))
                    )

                new_duree = edit_cols[2].text_input(
                    "DURÉE",
                    value=duree,
                    key=f"edit_duree_{idx}",
                )
                if st.button("Appliquer et décaler la suite", key=f"apply_shift_{idx}"):
                    updated = apply_timing_and_shift_following(
                        st.session_state["video_selection_df"],
                        idx,
                        new_time_in,
                        new_time_out,
                    )

                    st.session_state["video_selection_df"] = updated
                    st.session_state["video_detected_df"] = updated.copy()
                    st.rerun()

                delete_row = edit_cols[3].checkbox(
                    "Supprimer",
                    value=not bool(row.get("garder", True)),
                    key=f"delete_row_{idx}",
                )

                new_start_sec = tc_to_seconds(new_time_in)
                new_duration_sec = max(1, tc_to_seconds(new_duree))
                new_end_sec = new_start_sec + new_duration_sec

                edited = recompute_following_timings(
                    edited,
                    idx,
                    new_time_in,
                    new_time_out,
                )

                edited.at[idx, "garder"] = not delete_row

                st.caption(f"Détection ACR : {source_title} / durée {new_duree}")

                if st.button("📍 Aller à ce passage vidéo", key=f"go_video_{idx}"):
                    st.session_state["video_start_sec"] = new_start_sec
                    st.rerun()

                video_path_for_part = st.session_state.get("video_current_path", "")
                audio_excerpt = str(row.get("audio_excerpt", "")).strip()

                if video_path_for_part and Path(video_path_for_part).exists():
                    try:
                        audio_excerpt = make_audio_excerpt(
                            video_path_for_part,
                            new_start_sec,
                            new_duration_sec,
                            tempfile.gettempdir(),
                        )
                    except Exception as e:
                        log(f"Erreur extrait manuel {titre}: {repr(e)}")

                col1, col2 = st.columns(2)

                with col1:
                    st.markdown("#### Son détecté dans la vidéo")
                    if audio_excerpt and Path(audio_excerpt).exists():
                        st.audio(audio_excerpt)
                    else:
                        st.warning("Extrait audio indisponible.")

                with col2:
                    st.markdown("#### Master / référence")
                    if audio_ref and Path(audio_ref).exists():
                        st.audio(audio_ref)
                        st.caption(audio_ref)
                    else:
                        st.info("Titre commercial détecté")

                        yt_query = f"{titre} {artiste} official audio".replace(" ", "+")
                        spotify_query = f"{titre} {artiste}".replace(" ", "%20")

                        st.link_button(
                            "▶️ Écouter sur YouTube",
                            f"https://www.youtube.com/results?search_query={yt_query}",
                        )

                        st.link_button(
                            "🎵 Chercher sur Spotify",
                            f"https://open.spotify.com/search/{spotify_query}",
                        )

                        if isrc:
                            st.caption(f"ISRC : {isrc}")

                if str(row.get("SACEM STATUT", "")).strip():
                    st.markdown("#### Métadonnées SACEM")
                    st.write("Statut :", row.get("SACEM STATUT", ""))
                    st.write("Auteur(s) :", row.get("AUTEUR(S) SACEM", ""))
                    st.write("Compositeur(s) :", row.get("COMPOSITEUR(S) SACEM", ""))
                    st.write("Éditeur(s) :", row.get("EDITEUR(S) SACEM", ""))
                    st.write("Sous-éditeur(s) :", row.get("SOUS-EDITEUR(S) SACEM", ""))
                    st.write("Interprète(s) :", row.get("INTERPRETE(S) SACEM", ""))

                    
                    sacem_url = str(row.get("URL SACEM", "")).strip()
                    sacem_search_url = str(row.get("URL RECHERCHE SACEM", "")).strip()

                    if sacem_url.startswith("https://repertoire.sacem.fr/detail-oeuvre/"):
                         st.link_button("Ouvrir la fiche SACEM", sacem_url)

                    if "repertoire.sacem.fr" in sacem_search_url and "/resultats" in sacem_search_url:
                        st.link_button("Ouvrir la recherche SACEM", sacem_search_url)
                           
                   

                      
        st.session_state["video_selection_df"] = edited
        st.session_state["video_detected_df"] = edited.copy()
    if st.button("Construire le tableau final", type="primary"):
        selection_df = st.session_state.get("video_selection_df", pd.DataFrame()).copy()

        if selection_df.empty:
            st.warning("Aucune détection.")
        else:
            if "garder" in selection_df.columns:
                final_df = selection_df[selection_df["garder"] == True].copy()
            else:
                final_df = selection_df.copy()

            cols = [
                "part",
                "TITRE",
                "ARTISTE",
                "TIME IN",
                "TIME OUT",
                "DUREE",
                "MATCH SCORE",
                "ACR SCORE",
                "SOURCE TITLE",
                "isrc",
                "SACEM STATUT",
                "AUTEUR(S) SACEM",
                "COMPOSITEUR(S) SACEM",
                "EDITEUR(S) SACEM",
                "SOUS-EDITEUR(S) SACEM",
                "INTERPRETE(S) SACEM",
                "URL SACEM",
            ]

            for c in cols:
                if c not in final_df.columns:
                    final_df[c] = ""

            final_df = final_df[cols].copy()
            final_df["TITRE"] = final_df["TITRE"].map(normalize_title_decl)
            final_df["ARTISTE"] = final_df["ARTISTE"].astype(str).str.upper()

            st.session_state["results_df"] = final_df
            st.success("Tableau final construit.")


# ============================================================
# EXPORT
# ============================================================
if "results_df" in st.session_state:
    st.markdown("---")
    st.subheader("Résultats finaux")

    df_results = st.session_state["results_df"].copy()

    st.dataframe(df_results, width="stretch")

    st.download_button(
    "Télécharger Excel format M6",
    data=to_m6_excel_bytes(df_results),
    file_name="releve_droits_auteur_M6.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

    st.download_button(
        "Télécharger Excel",
        data=to_excel_bytes(df_results, sheet_name="Detection"),
        file_name="music_detection_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if debug_enabled:
    debug_panel()