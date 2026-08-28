import os
import time
import hmac
import base64
import hashlib
import tempfile
import subprocess
from pathlib import Path
import json
import re
import pandas as pd
import requests
from dotenv import load_dotenv
from fastapi import FastAPI, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from mdp_conductor import get_mdp_conductor
from mdp_solver import solve_mdp_timeline
from sacem_agent import SacemAgent
import asyncio
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from fastapi.responses import StreamingResponse
from io import BytesIO
from pathlib import Path
import unicodedata
from datetime import datetime, timezone
import uuid
import json
import sqlite3
import traceback
from fastapi import BackgroundTasks
from urllib.parse import quote
from fastapi import FastAPI, UploadFile, File, Form
from fastapi import (
    FastAPI,
    UploadFile,
    File,
    Form,
    HTTPException,
)
from playwright.sync_api import sync_playwright
import shutil
from demucs.separate import main as demucs_main
import sys
from pathlib import Path
import platform

def get_app_config_dir():
    system = platform.system()

    # Windows
    if system == "Windows":
        return (
            Path(os.environ["LOCALAPPDATA"])
            / "Music Rights"
        )

    # Mac
    if system == "Darwin":
        return (
            Path.home()
            / "Library"
            / "Application Support"
            / "Music Rights"
        )

    # Linux
    return (
        Path.home()
        / ".local"
        / "share"
        / "Music Rights"
    )


APP_CONFIG_DIR = get_app_config_dir()

APP_CONFIG_DIR.mkdir(
    parents=True,
    exist_ok=True,
)

ENV_PATH = APP_CONFIG_DIR / ".env"
def get_tool_path(tool_name: str) -> str:
    """
    Trouve ffmpeg / ffprobe.

    - Desktop Windows packagé :
      utilise les exécutables embarqués.
    - Développement / Mac :
      utilise ceux installés sur le système.
    """

    exe_name = (
        f"{tool_name}.exe"
        if os.name == "nt"
        else tool_name
    )

    # Application PyInstaller
    if getattr(sys, "frozen", False):
        internal_dir = Path(sys._MEIPASS)

        candidates = [
            internal_dir / "tools" / exe_name,
            internal_dir / exe_name,
        ]

        for candidate in candidates:
            if candidate.exists():
                print(
                    f"{tool_name.upper()} PATH:",
                    candidate,
                )
                return str(candidate)

    # Installation système / développement
    system_path = shutil.which(exe_name)

    if system_path:
        return system_path

    raise FileNotFoundError(
        f"{exe_name} introuvable. "
        "Music Rights ne contient pas "
        "l'outil nécessaire."
    )


def get_ffmpeg_path() -> str:
    return get_tool_path("ffmpeg")


def get_ffprobe_path() -> str:
    return get_tool_path("ffprobe")

load_dotenv(
    dotenv_path=ENV_PATH
)

ACR_HOST = os.getenv(
    "ACR_HOST",
    "",
).strip()

ACR_ACCESS_KEY = os.getenv(
    "ACR_ACCESS_KEY",
    "",
).strip()

ACR_ACCESS_SECRET = os.getenv(
    "ACR_ACCESS_SECRET",
    "",
).strip()

PROJECTS_DB = Path("projects.sqlite3")
JOBS_DB = Path("jobs.sqlite3")
JOBS_DIR = Path("cache/jobs")
JOBS_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI()
SACEM_CACHE_DB = Path("sacem_cache.sqlite3")
def init_jobs_db():
    with sqlite3.connect(JOBS_DB) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS analysis_jobs (
                id TEXT PRIMARY KEY,
                status TEXT NOT NULL,
                progress INTEGER NOT NULL DEFAULT 0,
                current_chunk INTEGER NOT NULL DEFAULT 0,
                total_chunks INTEGER NOT NULL DEFAULT 0,
                message TEXT NOT NULL DEFAULT '',
                video_path TEXT NOT NULL DEFAULT '',
                mapping_path TEXT NOT NULL DEFAULT '',
                parameters_json TEXT NOT NULL DEFAULT '{}',
                result_json TEXT,
                error TEXT,
                cancel_requested INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )

        conn.commit()


init_jobs_db()
def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def create_job(
    video_path: str,
    mapping_path: str,
    parameters: dict,
) -> str:
    job_id = uuid.uuid4().hex
    now = utc_now()

    with sqlite3.connect(JOBS_DB) as conn:
        conn.execute(
            """
            INSERT INTO analysis_jobs (
                id,
                status,
                progress,
                current_chunk,
                total_chunks,
                message,
                video_path,
                mapping_path,
                parameters_json,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                job_id,
                "pending",
                0,
                0,
                0,
                "En attente",
                video_path,
                mapping_path,
                json.dumps(parameters, ensure_ascii=False),
                now,
                now,
            ),
        )

        conn.commit()

    return job_id


def update_job(job_id: str, **fields):
    if not fields:
        return

    fields["updated_at"] = utc_now()

    assignments = ", ".join(
        f"{column} = ?"
        for column in fields
    )

    values = list(fields.values())
    values.append(job_id)

    with sqlite3.connect(JOBS_DB) as conn:
        conn.execute(
            f"""
            UPDATE analysis_jobs
            SET {assignments}
            WHERE id = ?
            """,
            values,
        )

        conn.commit()


def get_job(job_id: str) -> dict | None:
    with sqlite3.connect(JOBS_DB) as conn:
        conn.row_factory = sqlite3.Row

        row = conn.execute(
            """
            SELECT *
            FROM analysis_jobs
            WHERE id = ?
            """,
            (job_id,),
        ).fetchone()

    if row is None:
        return None

    job = dict(row)

    if job.get("result_json"):
        job["result"] = json.loads(
            job["result_json"]
        )
    else:
        job["result"] = None

    job["parameters"] = json.loads(
        job.get("parameters_json") or "{}"
    )

    return job


def job_cancel_requested(job_id: str) -> bool:
    with sqlite3.connect(JOBS_DB) as conn:
        row = conn.execute(
            """
            SELECT cancel_requested
            FROM analysis_jobs
            WHERE id = ?
            """,
            (job_id,),
        ).fetchone()

    return bool(row and row[0])
def init_projects_db():
    with sqlite3.connect(PROJECTS_DB) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                mode TEXT NOT NULL,
                video_name TEXT,
                rows_json TEXT NOT NULL,
                metadata_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.commit()
init_projects_db()
def normalize_cache_value(value: str) -> str:
    value = str(value or "").strip()

    value = (
        unicodedata.normalize("NFKD", value)
        .encode("ascii", "ignore")
        .decode("ascii")
    )

    value = value.upper()
    value = re.sub(r"[^A-Z0-9]+", " ", value)

    return " ".join(value.split())

import re


def clean_title_for_sacem(title: str) -> str:
    title = str(title or "").strip()

    # Supprimer complètement les parenthèses
    # Exemple :
    # Another Love (Piano Version)
    # -> Another Love
    title = re.sub(
        r"\([^)]*\)",
        " ",
        title,
    )

    # Supprimer aussi les crochets
    title = re.sub(
        r"\[[^\]]*\]",
        " ",
        title,
    )

    # Supprimer les suffixes fréquents après un tiret
    title = re.sub(
        r"\s*-\s*"
        r"(single version|radio edit|"
        r"album version|remastered.*|"
        r"original version|edit|"
        r"piano version|acoustic|instrumental)"
        r"\s*$",
        "",
        title,
        flags=re.IGNORECASE,
    )

    # Nettoyer les espaces
    title = re.sub(
        r"\s+",
        " ",
        title,
    ).strip()

    return title

def build_sacem_cache_key(title: str, artist: str) -> str:
    normalized_title = normalize_cache_value(title)
    normalized_artist = normalize_cache_value(artist)

    return f"{normalized_title}||{normalized_artist}"

def init_sacem_cache():
    with sqlite3.connect(SACEM_CACHE_DB) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS sacem_cache (
                cache_key TEXT PRIMARY KEY,
                title_input TEXT NOT NULL,
                artist_input TEXT NOT NULL,
                normalized_title TEXT NOT NULL,
                normalized_artist TEXT NOT NULL,
                status TEXT NOT NULL,
                result_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )

        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_sacem_cache_title
            ON sacem_cache(normalized_title)
            """
        )

        conn.commit()

 
init_sacem_cache()
 


app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

CONDUCTOR_WEEKDAY = [
    "MDP 2025 GENERIQUE DEBUT",
    "MDP 2025 NAPPE DEBUT",
    "MDP 2025 JINGLE NEUTRE",
    "MDP NAPPE TALK",
    "MDP 2025 JINGLE NEUTRE",
    "MDP 2025 JEU REFLEXION",
    "MDP 2025 NAPPE CHRONO",
    "MDP 2025 JEU REFLEXION",
    "MDP 2025 NAPPE CHRONO",
    "MDP 2025 JEU REFLEXION",
    "MDP 2025 NAPPE CHRONO",
    "MDP 2025 JEU REFLEXION",
    "MDP 2025 NAPPE CHRONO",
    "MDP 2025 JEU REFLEXION",
    "MDP 2025 NAPPE CHRONO",
    "MDP 2025 NAPPE FIN NEUTRE",
]

CONDUCTOR_FRIDAY = [
    "MDP 2025 GENERIQUE DEBUT",
    "MDP 2025 NAPPE DEBUT",
    "MDP 2025 JINGLE NEUTRE",
    "MDP NAPPE TALK",
    "MDP 2025 JINGLE NEUTRE",
    "MDP 2025 NAPPE FINALE REFLEXION",
    "MDP 2025 NAPPE CHRONO FINALE",
    "MDP 2025 NAPPE FINALE REFLEXION",
    "MDP 2025 NAPPE CHRONO FINALE",
    "MDP 2025 NAPPE FINALE REFLEXION",
    "MDP 2025 NAPPE CHRONO FINALE",
    "MDP 2025 NAPPE FINALE REFLEXION",
    "MDP 2025 NAPPE CHRONO FINALE",
    "MDP 2025 NAPPE FINALE REFLEXION",
    "MDP 2025 FIN GAGNANTE",
]


class TimelineRequest(BaseModel):
    conductor_type: str


class ShiftRequest(BaseModel):
    rows: list
    index: int
    time_in: str
    time_out: str

def normalize_title(value):
    value = str(value or "").lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())

SACEM_CACHE_DB = Path("sacem_cache.sqlite3")

def get_sacem_cache(
    title: str,
    artist: str,
):
    cache_key = build_sacem_cache_key(
        title,
        artist,
    )

    with sqlite3.connect(
        SACEM_CACHE_DB
    ) as conn:

        row = conn.execute(
            """
            SELECT status, result_json
            FROM sacem_cache
            WHERE cache_key = ?
            """,
            (cache_key,),
        ).fetchone()

    if row is None:
        return None

    status = str(
        row[0] or ""
    ).strip()

    # Un ancien NOT_FOUND ne doit jamais
    # empêcher une nouvelle recherche
    if status != "found":
        print(
            "SACEM CACHE IGNORED:",
            title,
            artist,
            status,
        )

        return None

    try:
        return json.loads(row[1])

    except Exception:
        return None


def save_sacem_cache(
    title: str,
    artist: str,
    result: dict,
):
    status = str(result.get("status", "")).strip()

    # Ne pas mémoriser les erreurs, blocages ou résultats douteux.
    if status != "found":
        return

    cache_key = build_sacem_cache_key(title, artist)
    normalized_title = normalize_cache_value(title)
    normalized_artist = normalize_cache_value(artist)
    now = datetime.now(timezone.utc).isoformat()

    with sqlite3.connect(SACEM_CACHE_DB) as conn:
        conn.execute(
            """
            INSERT INTO sacem_cache (
                cache_key,
                title_input,
                artist_input,
                normalized_title,
                normalized_artist,
                status,
                result_json,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)

            ON CONFLICT(cache_key) DO UPDATE SET
                title_input = excluded.title_input,
                artist_input = excluded.artist_input,
                normalized_title = excluded.normalized_title,
                normalized_artist = excluded.normalized_artist,
                status = excluded.status,
                result_json = excluded.result_json,
                updated_at = excluded.updated_at
            """,
            (
                cache_key,
                title,
                artist,
                normalized_title,
                normalized_artist,
                status,
                json.dumps(result, ensure_ascii=False),
                now,
                now,
            ),
        )

        conn.commit()
 
 
 
 

def to_m6_excel_bytes(df: pd.DataFrame) -> bytes:
    bio = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Relevé"

    yellow = PatternFill("solid", fgColor="FFC000")
    blue = PatternFill("solid", fgColor="BDD7EE")
    bold = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:J1")
    ws["A1"] = "RELEVE DE DROITS D'AUTEUR"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = "(Merci de nous retourner obligatoirement ce document)"
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["B5"] = "Réalisateur :"
    ws["B5"].font = Font(color="FF0000", bold=True)
    ws["B6"] = "Chaîne : M6"
    ws["B7"] = "Code product (si connu) :"

    ws["I5"] = "Titre du programme :"
    ws["I5"].font = bold
    ws["I6"] = "N° déclinaison :"
    ws["I7"] = "Date de diffusion :"
    ws["I8"] = "Informations complémentaires :"
    ws["I8"].fill = yellow

    headers = [
        "Lien associé à la ligne (colonne B)",
        "TITRE",
        "N° de produit (si connu)",
        "GENRE SACEM",
        "DUREE\n(HH:MM:SS:ii)",
        "TC IN",
        "TC OUT",
        "AUTEUR(S)",
        "COMPOSITEUR(S)",
        "EDITEUR(S)",
        "SOUS-EDITEUR(S)",
        "ISWC",
    ]

    start_row = 10

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col_idx)
        cell.value = header
        cell.fill = yellow
        cell.font = bold
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        values = [
            "",
            row.get("title", ""),
            "",
            row.get("genre_sacem", "FDS-Fond sonore"),
            row.get("duration", ""),
            row.get("time_in", ""),
            row.get("time_out", ""),
            row.get("auteur", ""),
            row.get("compositeur", ""),
            row.get("editeur", ""),
            row.get("sous_editeur", ""),
            row.get("code_iswc", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(r_idx, col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if col_idx in [2, 4, 5, 6, 7]:
                cell.fill = blue

    widths = {
        "A": 32, "B": 70, "C": 16, "D": 18,
        "E": 18, "F": 16, "G": 16,
        "H": 35, "I": 35, "J": 35,
        "K": 35, "L": 20,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[10].height = 38

    wb.save(bio)
    return bio.getvalue()

def pick_col(df, names):
    cols = {normalize_title(c): c for c in df.columns}

    for name in names:
        n = normalize_title(name)
        if n in cols:
            return cols[n]

    for c in df.columns:
        cn = normalize_title(c)
        for name in names:
            if normalize_title(name) in cn:
                return c

    return None


def load_sacem_mapping(path):
    df = pd.read_excel(path)

    title_col = pick_col(df, ["titre", "title", "oeuvre", "œuvre"])
    comp_col = pick_col(df, ["compositeur", "auteur", "authors", "composer"])
    edit_col = pick_col(df, ["editeur", "éditeur", "publisher"])
    iswc_col = pick_col(df, ["iswc"])
    isrc_col = pick_col(df, ["isrc"])

    if not title_col:
        return {}

    mapping = {}

    for _, r in df.iterrows():
        key = normalize_title(r.get(title_col, ""))

        if not key:
            continue

        mapping[key] = {
            "compositeur": str(r.get(comp_col, "") if comp_col else ""),
            "editeur": str(r.get(edit_col, "") if edit_col else ""),
            "code_iswc": str(r.get(iswc_col, "") if iswc_col else ""),
            "code_isrc": str(r.get(isrc_col, "") if isrc_col else ""),
        }

    return mapping

def sec_to_timecode(sec: int) -> str:
    sec = max(0, int(sec))
    return f"{sec // 3600:02d}:{(sec % 3600) // 60:02d}:{sec % 60:02d}"


def tc_to_seconds(tc: str) -> int:
    try:
        h, m, s = str(tc).split(":")[:3]
        return int(h) * 3600 + int(m) * 60 + int(s)
    except Exception:
        return 0


def fixed_duration(title: str):
    t = str(title).upper()
    if "GENERIQUE DEBUT" in t:
        return 10
    if "JINGLE NEUTRE" in t:
        return 3
    return None


def get_conductor(conductor_type: str):
    return CONDUCTOR_FRIDAY if conductor_type == "Vendredi" else CONDUCTOR_WEEKDAY


def load_stats():
    p = Path("mdp_stats.xlsx")
    if p.exists():
        return pd.read_excel(p)
    return pd.DataFrame()

def separate_music_with_demucs(input_wav: str) -> str:
    out_dir = tempfile.mkdtemp(prefix="demucs_")

    demucs_main([
    "--two-stems",
    "vocals",
    "-o",
    out_dir,
    input_wav,
])
    

    stem_name = Path(input_wav).stem
    no_vocals = Path(out_dir) / "htdemucs" / stem_name / "no_vocals.wav"

    if no_vocals.exists():
        return str(no_vocals)

    return input_wav

def duration_from_stats(title: str, position: int, conductor_type: str) -> int:
    fixed = fixed_duration(title)
    if fixed is not None:
        return fixed

    stats = load_stats()
    if not stats.empty:
        s = stats[
            (stats["type"].astype(str) == conductor_type)
            & (stats["position"].astype(int) == position)
            & (stats["title"].astype(str).str.upper().str.strip() == title.upper().strip())
        ]
        if not s.empty:
            return int(s.iloc[0]["median"])

    return 10

def build_conductor_rows(conductor_type: str):
    conductor = get_mdp_conductor(conductor_type)
    rows = []
    cursor = 0

    for i, item in enumerate(conductor):
        title = item["title"]
        mode = item.get("duration_mode", "historical")

        if mode == "fixed":
            duration = int(item.get("duration", 10))
        else:
            duration = duration_from_stats(title, i + 1, conductor_type)

       
        artist = ""

        

        start = cursor
        end = start + duration

        rows.append({
            "index": i,
            "title": title,
            "artist": artist,
            "duration_mode": mode,
            "expected_next": item.get("expected_next", ""),
            "time_in": sec_to_timecode(start),
            "time_out": sec_to_timecode(end),
            "duration": sec_to_timecode(duration),
            "start_sec": start,
            "end_sec": end,
            "score": "",
            "source": "conducteur intelligent",
        })

        cursor = end

    return rows

def resolve_dynamic_durations(rows):
    out = [r.copy() for r in rows]

    for i, row in enumerate(out):
        if row.get("duration_mode") != "dynamic":
            continue

        if i < len(out) - 1:
            start = int(row["start_sec"])
            next_start = int(out[i + 1]["start_sec"])
            duration = max(1, next_start - start)

            row["end_sec"] = next_start
            row["time_out"] = sec_to_timecode(next_start)
            row["duration"] = sec_to_timecode(duration)

    return out
@app.post("/projects")
async def save_project(
    name: str = Form(...),
    mode: str = Form(...),
    video_name: str = Form(""),
    rows_json: str = Form(...),
    metadata_json: str = Form("{}"),
):
    now = datetime.now(timezone.utc).isoformat()

    with sqlite3.connect(PROJECTS_DB) as conn:
        cursor = conn.execute(
            """
            INSERT INTO projects (
                name,
                mode,
                video_name,
                rows_json,
                metadata_json,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                name,
                mode,
                video_name,
                rows_json,
                metadata_json,
                now,
                now,
            ),
        )

        conn.commit()
        project_id = cursor.lastrowid

    return {
        "status": "ok",
        "project_id": project_id,
    }
@app.get("/projects")
def list_projects():
    with sqlite3.connect(PROJECTS_DB) as conn:
        conn.row_factory = sqlite3.Row

        rows = conn.execute(
            """
            SELECT
                id,
                name,
                mode,
                video_name,
                metadata_json,
                created_at,
                updated_at
            FROM projects
            ORDER BY updated_at DESC
            """
        ).fetchall()

    projects = []

    for row in rows:
        project = dict(row)

        metadata = json.loads(
            project.pop("metadata_json") or "{}"
        )

        project["duration"] = metadata.get(
            "videoDuration",
            0,
        )

        project["segment_count"] = metadata.get(
            "segmentCount",
            0,
        )

        project["validated_count"] = metadata.get(
            "validatedCount",
            0,
        )

        project["review_count"] = metadata.get(
            "reviewCount",
            0,
        )

        project["missing_count"] = metadata.get(
            "missingCount",
            0,
        )

        project["quality_score"] = metadata.get(
            "qualityScore",
            0,
        )

        projects.append(project)

    return {"projects": projects}
@app.get("/projects/{project_id}")
def get_project(project_id: int):
    with sqlite3.connect(PROJECTS_DB) as conn:
        conn.row_factory = sqlite3.Row

        row = conn.execute(
            """
            SELECT *
            FROM projects
            WHERE id = ?
            """,
            (project_id,),
        ).fetchone()

    if row is None:
        return {
            "status": "not_found",
        }

    project = dict(row)

    project["rows"] = json.loads(
        project.pop("rows_json")
    )

    project["metadata"] = json.loads(
        project.pop("metadata_json")
    )

    return project

@app.put("/projects/{project_id}")
async def update_project(
    project_id: int,
    name: str = Form(...),
    mode: str = Form(...),
    video_name: str = Form(""),
    rows_json: str = Form(...),
    metadata_json: str = Form("{}"),
):
    now = datetime.now(timezone.utc).isoformat()

    with sqlite3.connect(PROJECTS_DB) as conn:
        conn.execute(
            """
            UPDATE projects
            SET
                name = ?,
                mode = ?,
                video_name = ?,
                rows_json = ?,
                metadata_json = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                name,
                mode,
                video_name,
                rows_json,
                metadata_json,
                now,
                project_id,
            ),
        )

        conn.commit()

    return {
        "status": "ok",
        "project_id": project_id,
    }



@app.post("/timeline")
def create_timeline(req: TimelineRequest):
    return {"rows": build_conductor_rows(req.conductor_type)}

@app.post("/analyze-mdp")
async def analyze_mdp(
    video: UploadFile = File(...),
    conductor_type: str = Form("Lundi-Jeudi"),
):
    suffix = Path(video.filename).suffix or ".mp4"
    tmp_video = tempfile.NamedTemporaryFile(delete=False, suffix=suffix).name

    with open(tmp_video, "wb") as f:
        f.write(await video.read())

    rows = build_conductor_rows(conductor_type)
    rows = solve_mdp_timeline(rows, tmp_video)
    try:
         os.remove(tmp_video)
    except Exception:
        pass
    return {"rows": rows}

def acr_sign(string_to_sign: bytes, secret: str) -> str:
    return base64.b64encode(
        hmac.new(secret.encode(), string_to_sign, hashlib.sha1).digest()
    ).decode()


def recognize_chunk(wav_path: str):
    http_uri = "/v1/identify"
    timestamp = str(int(time.time()))

    string_to_sign = f"POST\n{http_uri}\n{ACR_ACCESS_KEY}\naudio\n1\n{timestamp}"
    signature = acr_sign(string_to_sign.encode(), ACR_ACCESS_SECRET)

    data = {
        "access_key": ACR_ACCESS_KEY,
        "sample_bytes": os.path.getsize(wav_path),
        "timestamp": timestamp,
        "signature": signature,
        "data_type": "audio",
        "signature_version": "1",
    }

    with open(wav_path, "rb") as f:
        r = requests.post(
            f"https://{ACR_HOST}{http_uri}",
            data=data,
            files={"sample": f},
            timeout=(20, 40),
        )

    r.raise_for_status()
    return r.json()

def analyze_acr_chunk(
    chunk_path: str,
    start: int,
    chunk_duration: int = 6,
    clean_audio: bool = False,
):
    acr_input = chunk_path

    if clean_audio:
        acr_input = clean_chunk_with_demucs(
            chunk_path
        )

    res = recognize_chunk(acr_input)

    status = res.get("status", {})
    metadata = res.get("metadata", {})

    if status.get("code") != 0:
        return []

    custom = (
        metadata.get("custom_files", [])
        or []
    )

    music = (
        metadata.get("music", [])
        or []
    )

    item = None
    source = ""

    if custom:
        item = custom[0]
        source = "custom_files"

    elif music:
        item = music[0]
        source = "music"

    if not item:
        return []

    title = str(
        item.get("title", "")
    ).strip()

    if not title:
        return []

    artist = ""

    artists = item.get("artists") or []

    if artists and isinstance(artists, list):
        artist = str(
            artists[0].get("name", "")
        ).strip()

    score = int(
        item.get("score") or 0
    )

    if score < 70:
        return []

    return [
        {
            "title": title,
            "artist": artist,
            "source": source,
            "start_sec": start,
            "end_sec": start + chunk_duration,
            "score": score,
        }
    ]

def extract_audio(video_path: str, wav_path: str):
    subprocess.check_call([
        get_ffmpeg_path(), "-y",
        "-i", video_path,
        "-vn",
        "-ac", "1",
        "-ar", "44100",
        wav_path,
    ])


def make_chunk(wav_path: str, start: int, duration: int, out_path: str):
    subprocess.run([
        get_ffmpeg_path(), "-y",
        "-ss", str(start),
        "-i", wav_path,
        "-t", str(duration),
        "-ac", "1",
        "-ar", "44100",
        out_path,
    ], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


def get_duration(path: str) -> int:
    cmd = [
        get_ffprobe_path(), "-v", "error",
        "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1",
        path,
    ]
    return int(float(subprocess.check_output(cmd).decode().strip()))


def clean_acr_hits(hits):
    if not hits:
        return []

    hits = sorted(hits, key=lambda x: x["start_sec"])
    cleaned = []
    current = hits[0].copy()

    for h in hits[1:]:
        same_title = h["title"] == current["title"]
        close = h["start_sec"] <= current["end_sec"]

        if same_title and close:
            current["end_sec"] = max(current["end_sec"], h["end_sec"])
            current["score"] = max(
                int(current.get("score") or 0),
                int(h.get("score") or 0),
            )
        else:
            cleaned.append(current)
            current = h.copy()

    cleaned.append(current)
    return cleaned

def clean_chunk_with_demucs(chunk_path: str) -> str:
    out_dir = tempfile.mkdtemp(prefix="demucs_chunk_")

    try:
        demucs_main([
            "--two-stems",
            "vocals",
            "-o",
            out_dir,
            chunk_path,
        ])

    except Exception as exc:
        print("========== DEMUCS ERROR ==========")
        print(repr(exc))
        print("==================================")

        return chunk_path

     

    stem = Path(chunk_path).stem
    cleaned = Path(out_dir) / "htdemucs" / stem / "no_vocals.wav"

    if cleaned.exists():
        return str(cleaned)

    return chunk_path


def merge_same_title_hits(hits, max_gap=18):
    if not hits:
        return []

    hits = sorted(hits, key=lambda h: int(h["start_sec"]))
    merged = []

    for h in hits:
        title = str(h.get("title", "")).strip().lower()

        if not merged:
            merged.append(h)
            continue

        last = merged[-1]
        last_title = str(last.get("title", "")).strip().lower()
        gap = int(h["start_sec"]) - int(last["end_sec"])

        if title == last_title and gap <= max_gap:
            last["end_sec"] = max(int(last["end_sec"]), int(h["end_sec"]))
            last["time_out"] = sec_to_timecode(last["end_sec"])
            last["duration"] = sec_to_timecode(
                int(last["end_sec"]) - int(last["start_sec"])
            )
            last["score"] = max(int(last.get("score") or 0), int(h.get("score") or 0))
        else:
            merged.append(h)

    return merged

def remove_overlapping_hits(hits):
    hits = sorted(hits, key=lambda h: (int(h["start_sec"]), int(h["end_sec"])))

    kept = []

    for h in hits:
        h_start = int(h["start_sec"])
        h_end = int(h["end_sec"])
        h_score = int(h.get("score") or 0)
        h_unknown = str(h.get("title", "")).upper() == "MUSIQUE NON RECONNUE"

        overlaps = [
            k for k in kept
            if h_start < int(k["end_sec"]) and h_end > int(k["start_sec"])
        ]

        if not overlaps:
            kept.append(h)
            continue

        # Si le nouveau est "non reconnu" et qu'il overlap un vrai titre, on l'ignore
        if h_unknown:
            continue

        # Si le nouveau est reconnu, il remplace les "non reconnus" qui overlapent
        kept = [
            k for k in kept
            if not (
                h_start < int(k["end_sec"])
                and h_end > int(k["start_sec"])
                and str(k.get("title", "")).upper() == "MUSIQUE NON RECONNUE"
            )
        ]

        # S'il overlap un autre vrai titre, on garde celui qui a le meilleur score
        real_overlaps = [
            k for k in kept
            if h_start < int(k["end_sec"]) and h_end > int(k["start_sec"])
        ]

        if not real_overlaps:
            kept.append(h)
        else:
            best_existing = max(real_overlaps, key=lambda x: int(x.get("score") or 0))
            if h_score > int(best_existing.get("score") or 0):
                kept = [k for k in kept if k not in real_overlaps]
                kept.append(h)

    return sorted(kept, key=lambda h: h["start_sec"])

def merge_repeated_titles(rows):
    grouped = {}
    order = []

    for row in rows:
        title = str(row.get("title", "")).strip()
        artist = str(row.get("artist", "")).strip()

        key = (
            normalize_title(title),
            normalize_title(artist),
        )

        if key not in grouped:
            grouped[key] = {
                **row,
                "start_sec": int(row.get("start_sec", 0)),
                "end_sec": int(row.get("end_sec", 0)),
                "total_duration_sec": max(
                    0,
                    int(row.get("end_sec", 0))
                    - int(row.get("start_sec", 0)),
                ),
                "appearances": 1,
            }

            order.append(key)
            continue

        current = grouped[key]

        start = int(row.get("start_sec", 0))
        end = int(row.get("end_sec", 0))
        duration = max(0, end - start)

        current["start_sec"] = min(
            int(current["start_sec"]),
            start,
        )

        current["end_sec"] = max(
            int(current["end_sec"]),
            end,
        )

        current["total_duration_sec"] += duration
        current["appearances"] += 1

        current["time_in"] = sec_to_timecode(
            current["start_sec"]
        )

        current["time_out"] = sec_to_timecode(
            current["end_sec"]
        )

        current["duration"] = sec_to_timecode(
            current["total_duration_sec"]
        )

    merged = []

    for index, key in enumerate(order):
        row = grouped[key]
        row["index"] = index
        merged.append(row)

    return merged


def enrich_sacem_sync(rows: list) -> list:
    """
    Enrichit les lignes avec les informations SACEM.

    - Utilise le cache si disponible.
    - Nettoie le titre avant recherche SACEM.
    - Retente avec le titre seul si nécessaire.
    - Met en cache uniquement les résultats "found".
    - Conserve les anciennes données si la SACEM
      est bloquée ou rencontre une erreur.
    """

    is_server_linux = (
        platform.system() == "Linux"
    )

    agent = SacemAgent(
    headless=is_server_linux
)
    enriched = []

    for position, row in enumerate(rows):
        print("ROW =", row)

        # Copie de la ligne existante.
        # Important : cela permet de conserver les anciennes
        # informations SACEM si la nouvelle recherche échoue.
        new_row = dict(row)

        title = str(
            new_row.get("title", "")
        ).strip()

        artist = str(
            new_row.get("artist", "")
            or new_row.get("artiste", "")
        ).strip()

        # ----------------------------------------
        # 1. Titre vide
        # ----------------------------------------
        if not title:
            new_row["statut_sacem"] = "titre vide"
            enriched.append(new_row)
            continue

        # ----------------------------------------
        # 2. Titres internes à ignorer
        # ----------------------------------------
        internal_titles = (
            "GENERIQUE",
            "GÉNÉRIQUE",
            "JINGLE",
            "NAPPE",
            "MDP ",
        )

        if title.upper().startswith(
            internal_titles
        ):
            new_row["statut_sacem"] = (
                "ignoré - titre interne"
            )

            enriched.append(new_row)
            continue

        try:
            # ----------------------------------------
            # 3. Recherche dans le cache
            # ----------------------------------------
            res = get_sacem_cache(
                title,
                artist,
            )

            if res is not None:
                print(
                    "SACEM CACHE HIT:",
                    title,
                    artist,
                )

                source_sacem = "cache"

            else:
                print(
                    "SACEM CACHE MISS:",
                    title,
                    artist,
                )

                source_sacem = "sacem"

                # ----------------------------------------
                # 4. Nettoyage du titre
                # ----------------------------------------
                sacem_title = clean_title_for_sacem(
                    title
                )

                print(
                    "SACEM TITLE:",
                    repr(title),
                    "->",
                    repr(sacem_title),
                )

                # ----------------------------------------
                # 5. Recherche :
                # titre nettoyé + artiste
                # ----------------------------------------
                res = agent.search(
                    sacem_title,
                    artist,
                )

                # ----------------------------------------
                # 6. Si pas trouvé :
                # titre nettoyé seul
                # ----------------------------------------
                if (
                    res.get("status")
                    == "not_found"
                ):
                    print(
                        "SACEM RETRY TITLE ONLY:",
                        sacem_title,
                    )

                    res = agent.search(
                        sacem_title,
                        "",
                    )

                # ----------------------------------------
                # 7. Cache uniquement les FOUND
                # ----------------------------------------
                if (
                    res.get("status")
                    == "found"
                ):
                    save_sacem_cache(
                        title,
                        artist,
                        res,
                    )

            print(
                "SACEM RESULT:",
                res,
            )

            print(
                "SACEM SOURCE:",
                source_sacem,
            )

            # ----------------------------------------
            # 8. SACEM BLOQUÉE
            # ----------------------------------------
            if (
                res.get("status")
                == "blocked"
            ):
                print(
                    "SACEM BLOQUEE - "
                    "conservation des anciennes données"
                )

                # Vérifier si cette ligne possédait
                # déjà des informations SACEM.
                has_old_sacem_data = bool(
                    new_row.get("compositeur")
                    or new_row.get("auteur")
                    or new_row.get("editeur")
                    or new_row.get("code_iswc")
                    or new_row.get(
                        "url_sacem_detail"
                    )
                    or new_row.get(
                        "url_sacem_candidate"
                    )
                )

                if has_old_sacem_data:
                    # --------------------------------
                    # NE PAS écraser les anciennes infos
                    # --------------------------------

                    old_status = str(
                        row.get(
                            "statut_sacem",
                            "",
                        )
                    ).strip()

                    # Si l'ancien résultat était trouvé,
                    # on conserve FOUND.
                    if old_status == "found":
                        new_row[
                            "statut_sacem"
                        ] = "found"

                    # Sinon on conserve son ancien statut.
                    elif old_status:
                        new_row[
                            "statut_sacem"
                        ] = old_status

                    else:
                        new_row[
                            "statut_sacem"
                        ] = "à vérifier"

                    new_row[
                        "source_sacem"
                    ] = (
                        row.get(
                            "source_sacem"
                        )
                        or
                        "ancien résultat conservé"
                    )

                else:
                    # Aucun ancien résultat disponible.
                    new_row[
                        "statut_sacem"
                    ] = "blocked"

                    new_row[
                        "source_sacem"
                    ] = "sacem"

                enriched.append(new_row)

                # SACEM vient de bloquer l'accès.
                # On arrête immédiatement les recherches.
                #
                # Les autres lignes sont conservées
                # EXACTEMENT telles qu'elles étaient.
                for remaining in (
                    rows[position + 1:]
                ):
                    enriched.append(
                        dict(remaining)
                    )

                break

            # ----------------------------------------
            # 9. Copier le nouveau résultat SACEM
            # ----------------------------------------

            new_row[
                "statut_sacem"
            ] = res.get(
                "status",
                "",
            )

            new_row[
                "compositeur"
            ] = "; ".join(
                res.get(
                    "composers",
                    [],
                )
            )

            new_row[
                "auteur"
            ] = "; ".join(
                res.get(
                    "authors",
                    [],
                )
            )

            new_row[
                "editeur"
            ] = "; ".join(
                res.get(
                    "publishers",
                    [],
                )
            )

            new_row[
                "sous_editeur"
            ] = "; ".join(
                res.get(
                    "sub_publishers",
                    [],
                )
            )

            new_row[
                "interprete"
            ] = "; ".join(
                res.get(
                    "performers",
                    [],
                )
            )

            new_row[
                "code_iswc"
            ] = res.get(
                "iswc",
                "",
            )

            # ----------------------------------------
            # 10. URLs SACEM
            # ----------------------------------------

            # Fiche SACEM acceptée
            new_row[
                "url_sacem_detail"
            ] = res.get(
                "url",
                "",
            )

            # Fiche candidate :
            # résultat trouvé mais matching insuffisant
            new_row[
                "url_sacem_candidate"
            ] = res.get(
                "candidate_url",
                "",
            )

            # URL de la recherche SACEM
            search_url = res.get(
                "search_url",
                "",
            )

            # Si l'agent n'a pas retourné
            # l'URL de recherche, on la construit.
            if not search_url:
                query = title

                if artist:
                    query = (
                        f"{title},{artist}"
                    )

                search_url = (
                    "https://www."
                    "repertoire.sacem.fr/"
                    "resultats?"
                    "filters=titles,parties"
                    f"&query={quote(query)}"
                    "#searchBtn"
                )

            new_row[
                "url_sacem"
            ] = search_url

            new_row[
                "source_sacem"
            ] = source_sacem

        # ----------------------------------------
        # 11. ERREUR SACEM
        # ----------------------------------------
        except Exception as exc:
            print(
                "SACEM ERROR:",
                repr(exc),
            )

            # Vérifier si nous possédions déjà
            # des informations SACEM.
            has_old_sacem_data = bool(
                new_row.get("compositeur")
                or new_row.get("auteur")
                or new_row.get("editeur")
                or new_row.get("code_iswc")
                or new_row.get(
                    "url_sacem_detail"
                )
                or new_row.get(
                    "url_sacem_candidate"
                )
            )

            if has_old_sacem_data:
                # --------------------------------
                # Conserver les anciennes données
                # --------------------------------
                print(
                    "Anciennes données SACEM "
                    "conservées"
                )

                new_row[
                    "source_sacem"
                ] = (
                    new_row.get(
                        "source_sacem"
                    )
                    or
                    "ancien résultat conservé"
                )

            else:
                # --------------------------------
                # Aucun ancien résultat
                # --------------------------------
                new_row[
                    "statut_sacem"
                ] = (
                    f"error: "
                    f"{type(exc).__name__}: "
                    f"{exc}"
                )

                new_row[
                    "url_sacem"
                ] = ""

                new_row[
                    "url_sacem_detail"
                ] = ""

                new_row[
                    "url_sacem_candidate"
                ] = ""

        enriched.append(new_row)

    return enriched

@app.post("/enrich-sacem")
async def enrich_sacem(
    rows_json: str = Form(...),
    mapping: UploadFile | None = File(None),
):
    rows = json.loads(rows_json)

    print("ROWS RECEIVED:", len(rows), flush=True)

    enriched = await asyncio.to_thread(
        enrich_sacem_sync,
        rows,
    )

    # garde ici la suite de ton code actuel

    return {"rows": enriched}

@app.get("/sacem-cache/stats")
def sacem_cache_stats():
    with sqlite3.connect(SACEM_CACHE_DB) as conn:
        total = conn.execute(
            "SELECT COUNT(*) FROM sacem_cache"
        ).fetchone()[0]

        found = conn.execute(
            """
            SELECT COUNT(*)
            FROM sacem_cache
            WHERE status = 'found'
            """
        ).fetchone()[0]

    return {
        "total": total,
        "found": found,
    }


@app.delete("/sacem-cache")
def clear_sacem_cache():
    with sqlite3.connect(SACEM_CACHE_DB) as conn:
        deleted = conn.execute(
            "DELETE FROM sacem_cache"
        ).rowcount

        conn.commit()

    return {
        "status": "ok",
        "deleted": deleted,
    }

GENERIC_RULES = [
    {
        "title": "GENERIQUE TALENT",
        "max_start": 25,
        "keywords": [
            "generique talent",
            "talent intro",
        ],
        "fixed_duration": 17,
    },
    {
        "title": "GENERIQUE ADP",
        "max_start": 25,
        "keywords": [
            "generique adp",
            "adp intro",
        ],
        "fixed_duration": 17,
    },
]
def detect_generic(rows, intro_type="AUTO"):
    if intro_type != "AUTO":
        return rows

    for rule in GENERIC_RULES:
        for row in rows:
            start = int(row.get("start_sec", 0))
            acr_title = normalize_title(
                row.get("acr_title")
                or row.get("title")
            )

            if start > rule["max_start"]:
                continue

            if any(
                keyword in acr_title
                for keyword in rule["keywords"]
            ):
                duration = int(rule["fixed_duration"])

                generic = {
                    "index": 0,
                    "title": rule["title"],
                    "artist": "",
                    "acr_title": row.get("acr_title", ""),
                    "time_in": "00:00:00",
                    "time_out": sec_to_timecode(duration),
                    "duration": sec_to_timecode(duration),
                    "start_sec": 0,
                    "end_sec": duration,
                    "score": row.get("score", ""),
                    "source": "générique automatique",
                }

                remaining = [
                    item
                    for item in rows
                    if int(item.get("start_sec", 0)) >= duration
                ]

                result = [generic, *remaining]

                for index, item in enumerate(result):
                    item["index"] = index

                return result

    return rows

 


@app.post("/analyze-acr")
async def analyze_acr(
    video: UploadFile = File(...),
    conductor_type: str = Form("Lundi-Jeudi"),
    mapping: UploadFile | None = File(None),
    intro_type: str = Form("NONE"),
    clean_audio: str = Form("NO"),
):
    tmp_video = tempfile.NamedTemporaryFile(delete=False, suffix=".mp4").name

    with open(tmp_video, "wb") as f:
        f.write(await video.read())
    print("VIDEO SIZE:", os.path.getsize(tmp_video))

    mapping_path = None
    if mapping is not None and mapping.filename:
        suffix = Path(mapping.filename).suffix or ".xlsx"
        mapping_path = tempfile.NamedTemporaryFile(delete=False, suffix=suffix).name
        with open(mapping_path, "wb") as f:
            f.write(await mapping.read())
        print("MAPPING FILE:", mapping_path)

     #tmp_wav = tempfile.NamedTemporaryFile(delete=False, suffix=".wav").name
     #extract_audio(tmp_video, tmp_wav)
     # tmp_wav = separate_music_with_demucs(tmp_wav)
    total_duration = get_duration(tmp_video)

 
    error_count = 0
    max_errors = 15
    chunk_duration = 6
    step = 12
    hits = []
     
    for start in range(0, total_duration, step):

        chunk_path = tempfile.NamedTemporaryFile(delete=False, suffix=".wav").name
        make_chunk(tmp_video, start, chunk_duration, chunk_path)
        acr_input = chunk_path

        if clean_audio == "YES":
            acr_input = clean_chunk_with_demucs(chunk_path)

        try:
            res = recognize_chunk(acr_input)
            time.sleep(1)
            status = res.get("status", {})
            metadata = res.get("metadata", {})

            if status.get("code") != 0:
                
                continue

            custom = metadata.get("custom_files", []) or []
            music = metadata.get("music", []) or []

            item = None
            source = ""

            if custom:
                item = custom[0]
                source = "custom_files"
            elif music:
                item = music[0]
                source = "music"

            if not item:
                 
                continue

            title = str(item.get("title", "")).strip()
            artist = ""

            artists = item.get("artists") or []
            if artists and isinstance(artists, list):
                artist = str(artists[0].get("name", "")).strip()
            if not title:
                continue
            score = int(item.get("score") or 0)

            if score < 70:
                 continue
          

            hits.append({
                "title": title,
                "artist": artist,
                "source": source,
                "start_sec": start,
                "end_sec": start + chunk_duration,
                "score": score,
            })

        except Exception as e:
            error_count += 1
            print("ACR error:", repr(e))

            if error_count >= max_errors:
                print("ACR STOP: trop d'erreurs ACRCloud")
                break
        finally:
            try:
                os.remove(chunk_path)
            except Exception:
                pass
    print("CHUNK:", start, "/", total_duration)
    raw_hits = hits
    cleaned_hits = clean_acr_hits(hits)
    cleaned_hits = remove_overlapping_hits(cleaned_hits)
    cleaned_hits = merge_same_title_hits(cleaned_hits)
 

    print("ACR RAW COUNT:", len(raw_hits))
    print("ACR CLEANED COUNT:", len(cleaned_hits))
    print("CLEANED HITS:", cleaned_hits[:10])
    mapping_dict = load_mapping_titles(mapping_path)
    acr_rows = []
     
    if intro_type == "GEN_TALENT":
        acr_rows.append({
        "index": 0,
        "title": "GENERIQUE TALENT",
        "acr_title": "",
        "time_in": "00:00:00",
        "time_out": "00:00:17",
        "duration": "00:00:17",
        "start_sec": 0,
        "end_sec": 17,
        "score": "",
        "source": "intro fixe",
    })

    elif intro_type == "GEN_ADP":
        acr_rows.append({
        "index": 0,
        "title": "GENERIQUE ADP",
        "artist": "",
        "acr_title": "",
        "time_in": "00:00:00",
        "time_out": "00:00:17",
        "duration": "00:00:17",
        "start_sec": 0,
        "end_sec": 17,
        "score": "",
        "source": "intro fixe",
    })
 
    for i, h in enumerate(cleaned_hits):
        start = int(h["start_sec"])
        end = int(h["end_sec"])

        acr_rows.append({
            "index": len(acr_rows),
            "title": map_acr_title(h["title"], mapping_dict),
            "artist": h.get("artist", ""),
            "acr_title": h["title"],
            "time_in": sec_to_timecode(start),
            "time_out": sec_to_timecode(end),
            "duration": sec_to_timecode(end - start),
            "start_sec": start,
            "end_sec": end,
            "score": h.get("score", ""),
            "source": "ACRCloud seul",
        })
    acr_rows = merge_repeated_titles(acr_rows)

    acr_rows = detect_generic(
    acr_rows,
    intro_type=intro_type,
)
    acr_rows = merge_repeated_titles(acr_rows)
    try:
         os.remove(tmp_video)
    except Exception:
        pass

    if mapping_path:
        try:
             os.remove(mapping_path)
        except Exception:
            pass

    return {
        "rows": acr_rows,
        "acr_hits": cleaned_hits,
        "acr_raw_hits": raw_hits,
        "video_duration": total_duration,
    }

def project_excel_bytes(rows, metadata=None):
    metadata = metadata or {}

    bio = BytesIO()
    wb = Workbook()

    ws = wb.active
    ws.title = "Emission"

    project_headers = [
        "Index",
        "Titre",
        "Artiste",
        "TC IN",
        "TC OUT",
        "Durée cumulée",
        "Début secondes",
        "Fin secondes",
        "Apparitions",
        "Score ACR",
        "Source ACR",
        "Statut SACEM",
        "Titre SACEM",
        "Auteur",
        "Compositeur",
        "Editeur",
        "Sous-éditeur",
        "Interprète",
        "ISWC",
        "URL SACEM",
        "Source SACEM",
    ]

    for column, header in enumerate(
        project_headers,
        start=1,
    ):
        cell = ws.cell(1, column)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            "solid",
            fgColor="FFC000",
        )

    for row_index, row in enumerate(rows, start=2):
        values = [
            row.get("index", ""),
            row.get("title", ""),
            row.get("artist", ""),
            row.get("time_in", ""),
            row.get("time_out", ""),
            row.get("duration", ""),
            row.get("start_sec", ""),
            row.get("end_sec", ""),
            row.get("appearances", 1),
            row.get("score", ""),
            row.get("source", ""),
            row.get("statut_sacem", ""),
            row.get("titre_sacem", ""),
            row.get("auteur", ""),
            row.get("compositeur", ""),
            row.get("editeur", ""),
            row.get("sous_editeur", ""),
            row.get("interprete", ""),
            row.get("code_iswc", ""),
            row.get("url_sacem", ""),
            row.get("source_sacem", ""),
        ]

        for column, value in enumerate(values, start=1):
            ws.cell(row_index, column).value = value

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    info = wb.create_sheet("Informations")
    info.append(["Champ", "Valeur"])

    for key, value in metadata.items():
        info.append([key, value])

    wb.save(bio)
    return bio.getvalue()
@app.post("/save-project-excel")
async def save_project_excel(
    rows_json: str = Form(...),
    metadata_json: str = Form("{}"),
):
    rows = json.loads(rows_json)
    metadata = json.loads(metadata_json)

    excel = project_excel_bytes(
        rows,
        metadata,
    )

    return StreamingResponse(
        BytesIO(excel),
        media_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                'attachment; filename="Projet_Emission.xlsx"'
            )
        },
    )
def build_acr_rows(
    cleaned_hits,
    mapping_path=None,
    intro_type="NONE",
):
    mapping_dict = load_mapping_titles(
        mapping_path
    )

    acr_rows = []

    # Générique fixe éventuel
    if intro_type == "GEN_TALENT":
        acr_rows.append({
            "index": 0,
            "title": "GENERIQUE TALENT",
            "artist": "",
            "acr_title": "",
            "time_in": "00:00:00",
            "time_out": "00:00:17",
            "duration": "00:00:17",
            "start_sec": 0,
            "end_sec": 17,
            "score": "",
            "source": "intro fixe",
        })

    elif intro_type == "GEN_ADP":
        acr_rows.append({
            "index": 0,
            "title": "GENERIQUE ADP",
            "artist": "",
            "acr_title": "",
            "time_in": "00:00:00",
            "time_out": "00:00:17",
            "duration": "00:00:17",
            "start_sec": 0,
            "end_sec": 17,
            "score": "",
            "source": "intro fixe",
        })

    # Résultats ACRCloud
    for hit in cleaned_hits:
        start = int(
            hit.get("start_sec", 0)
        )

        end = int(
            hit.get("end_sec", start + 1)
        )

        if end <= start:
            end = start + 1

        original_title = str(
            hit.get("title", "")
        ).strip()

        final_title = map_acr_title(
            original_title,
            mapping_dict,
        )

        acr_rows.append({
            "index": len(acr_rows),
            "title": final_title,
            "artist": str(
                hit.get("artist", "")
            ).strip(),
            "acr_title": original_title,
            "time_in": sec_to_timecode(start),
            "time_out": sec_to_timecode(end),
            "duration": sec_to_timecode(
                end - start
            ),
            "start_sec": start,
            "end_sec": end,
            "score": hit.get("score", ""),
            "source": "ACRCloud seul",
        })

    # Fusion des répétitions
    acr_rows = merge_repeated_titles(
        acr_rows
    )

    # Détection éventuelle du générique AUTO
    acr_rows = detect_generic(
        acr_rows,
        intro_type=intro_type,
    )

    # On refusionne après la détection
    acr_rows = merge_repeated_titles(
        acr_rows
    )

    # Réindexation propre
    for index, row in enumerate(acr_rows):
        row["index"] = index

    return acr_rows

def run_acr_job(job_id: str):
    job = get_job(job_id)

    if job is None:
        return

    try:
        update_job(
            job_id,
            status="running",
            message="Lecture de la vidéo",
            progress=1,
        )

        video_path = job["video_path"]
        mapping_path = job["mapping_path"]
        parameters = job["parameters"]

        total_duration = get_duration(video_path)

        step = 12

        chunk_starts = list(
            range(
                0,
                int(total_duration),
                step,
            )
        )

        total_chunks = len(chunk_starts)

        update_job(
            job_id,
            total_chunks=total_chunks,
            message="Analyse ACRCloud",
            progress=2,
        )

        checkpoint_path = (
            JOBS_DIR / job_id / "checkpoint.json"
        )

        completed_chunks = 0
        hits = []

        if checkpoint_path.exists():
            checkpoint = json.loads(
                checkpoint_path.read_text(
                    encoding="utf-8"
                )
            )

            completed_chunks = int(
                checkpoint.get(
                    "completed_chunks",
                    0,
                )
            )

            hits = checkpoint.get("hits", [])

        for chunk_index, start in enumerate(
            chunk_starts,
            start=1,
        ):
            if chunk_index <= completed_chunks:
                continue

            if job_cancel_requested(job_id):
                update_job(
                    job_id,
                    status="cancelled",
                    message="Analyse interrompue",
                )
                return

            chunk_path = (
                JOBS_DIR
                / job_id
                / f"chunk_{chunk_index:05d}.wav"
            )

            make_chunk(
                video_path,
                start,
                step,
                str(chunk_path),
            )

            chunk_hits = analyze_acr_chunk(
                str(chunk_path),
                start,
            )

            hits.extend(chunk_hits)

            progress = round(
                chunk_index
                / max(total_chunks, 1)
                * 85
            )

            update_job(
                job_id,
                progress=progress,
                current_chunk=chunk_index,
                total_chunks=total_chunks,
                message=(
                    f"Analyse ACRCloud "
                    f"{chunk_index}/{total_chunks}"
                ),
            )

            checkpoint_path.write_text(
                json.dumps(
                    {
                        "completed_chunks":
                            chunk_index,
                        "hits": hits,
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            try:
                chunk_path.unlink()
            except OSError:
                pass

        update_job(
            job_id,
            progress=88,
            message="Nettoyage des résultats",
        )

        cleaned_hits = clean_acr_hits(hits)
        cleaned_hits = remove_overlapping_hits(
            cleaned_hits
        )
        cleaned_hits = merge_same_title_hits(
            cleaned_hits
        )

        update_job(
            job_id,
            progress=93,
            message="Création de la timeline",
        )

        rows = build_acr_rows(
            cleaned_hits=cleaned_hits,
            mapping_path=mapping_path,
            intro_type=parameters.get(
                "intro_type",
                "NONE",
            ),
        )

        result = {
            "rows": rows,
            "acr_hits": cleaned_hits,
            "video_duration": total_duration,
        }

        result_path = (
            JOBS_DIR / job_id / "result.json"
        )

        result_path.write_text(
            json.dumps(
                result,
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        update_job(
            job_id,
            status="done",
            progress=100,
            current_chunk=total_chunks,
            message="Analyse terminée",
            result_json=json.dumps(
                result,
                ensure_ascii=False,
            ),
        )

    except Exception as exc:
        error_text = (
            f"{type(exc).__name__}: {exc}\n"
            f"{traceback.format_exc()}"
        )

        update_job(
            job_id,
            status="error",
            message="Erreur pendant l’analyse",
            error=error_text,
        )

def load_mapping_titles(mapping_path):
    if not mapping_path:
        return {}

    try:
        df = pd.read_excel(mapping_path)
    except Exception:
        return {}

    mapping = {}

    # Col A = titre final, Col B = intitulé outil
    for _, row in df.iterrows():
        final_title = str(row.iloc[0]).strip()
        tool_title = str(row.iloc[1]).strip()

        if final_title and tool_title and final_title.lower() != "nan":
            mapping[tool_title.upper()] = final_title

    return mapping


def map_acr_title(acr_title, mapping_dict):
    t = str(acr_title).upper().strip()

    for tool_title, final_title in mapping_dict.items():
        if tool_title in t or t in tool_title:
            return final_title

    return acr_title
@app.post("/shift")
def shift_rows(req: ShiftRequest):
    rows = req.rows

    start = tc_to_seconds(req.time_in)
    end = tc_to_seconds(req.time_out)

    if end <= start:
        end = start + 1

    rows[req.index]["start_sec"] = start
    rows[req.index]["end_sec"] = end
    rows[req.index]["time_in"] = sec_to_timecode(start)
    rows[req.index]["time_out"] = sec_to_timecode(end)
    rows[req.index]["duration"] = sec_to_timecode(end - start)

    cursor = end

    for i in range(req.index + 1, len(rows)):
        old_duration = max(1, int(rows[i]["end_sec"]) - int(rows[i]["start_sec"]))

        rows[i]["start_sec"] = cursor
        rows[i]["end_sec"] = cursor + old_duration
        rows[i]["time_in"] = sec_to_timecode(cursor)
        rows[i]["time_out"] = sec_to_timecode(cursor + old_duration)
        rows[i]["duration"] = sec_to_timecode(old_duration)

        cursor += old_duration

    return {"rows": rows}
 
def search(self, title: str, artist: str = "") -> dict:
    title = str(title).strip()
    artist = str(artist).strip()

    with sync_playwright() as p:
                

        CHROMIUM_PATHS = [
            "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
            "/Applications/Chromium.app/Contents/MacOS/Chromium",
            "/opt/homebrew/bin/chromium",
        ]

        chromium_executable = next(
            (
                path
                for path in CHROMIUM_PATHS
                if os.path.exists(path)
            ),
            None,
        )

        if chromium_executable:
            browser = p.chromium.launch(
                headless=True,
                executable_path=chromium_executable,
            )
        else:
            browser = p.chromium.launch(
                headless=True,
            )

        context = browser.new_context(
        viewport={"width": 1450, "height": 900},
        locale="fr-FR",
    )

        page = context.new_page()

        try:
            body_text = self._run_search(page, title, artist)

            if self._is_blocked(body_text):
                return {
                    "status": "blocked",
                    "title_input": title,
                    "artist_input": artist,
                    "search_mode": "title_artist",
                    "url": page.url,
                    "raw_text": body_text[:2000],
                }

            search_url = page.url
            search_mode = "title_artist"

            if not self._has_results(body_text):
                body_text = self._run_search(page, title, "")

                if self._is_blocked(body_text):
                    return {
                        "status": "blocked",
                        "title_input": title,
                        "artist_input": artist,
                        "search_mode": "title_only_fallback",
                        "url": page.url,
                        "raw_text": body_text[:2000],
                    }

                search_mode = "title_only_fallback"
                search_url = page.url

            detail_links = page.get_by_text(
                "VOIR LE DÉTAIL",
                exact=True,
            )

            count = detail_links.count()

            if count == 0:
                return {
                    "status": "not_found",
                    "title_input": title,
                    "artist_input": artist,
                    "search_mode": search_mode,
                    "url": "",
                    "raw_text": body_text[:2000],
                }

            selected_index = self._choose_result_index_from_page(
                page,
                title,
                artist if search_mode == "title_artist" else "",
            )

            if selected_index >= count:
                selected_index = 0

            selected_link = detail_links.nth(selected_index)

            selected_link.scroll_into_view_if_needed()

            selected_link.click(
                timeout=10000,
                force=True,
            )

            try:
                page.wait_for_url(
                    "**/detail-oeuvre/**",
                    timeout=10000,
                )
            except Exception:
                pass

            page.wait_for_timeout(2000)

            current_url = page.url

            if "/detail-oeuvre/" not in current_url:
                return {
                    "status": "not_found",
                    "title": "",
                    "iswc": "",
                    "authors": [],
                    "composers": [],
                    "publishers": [],
                    "sub_publishers": [],
                    "performers": [],
                    "url": "",
                    "artist_input": artist,
                    "title_input": title,
                    "search_mode": search_mode,
                    "result_count": count,
                    "selected_result_index": selected_index,
                    "search_url": search_url,
                }

            current_title = page.title()
            detail_text = page.locator("body").inner_text()

            parsed = parse_sacem_detail(detail_text)

            returned_title = parsed.get("title", "")
            title_score = fuzz.ratio(
                normalize_key(title),
                normalize_key(returned_title),
            )

            parsed["title_match_score"] = title_score

            if title_score < 85:
                parsed["status"] = "not_found"
                parsed["candidate_url"] = current_url
                parsed["url"] = ""

            else:
                parsed["status"] = "found"
                parsed["url"] = current_url
                parsed["candidate_url"] = ""

            print(
                "SACEM DETAIL URL:",
                current_url,
            )

            print(
                "SACEM CANDIDATE URL:",
                parsed.get("candidate_url", ""),
            )
            
            parsed["artist_input"] = artist
            parsed["title_input"] = title
            parsed["search_mode"] = search_mode
            parsed["page_title"] = current_title
            parsed["result_count"] = count
            parsed["selected_result_index"] = selected_index
            parsed["search_url"] = search_url

            return parsed

        finally:
            browser.close()

@app.post("/download-m6")
async def download_m6(rows_json: str = Form(...)):
    rows = json.loads(rows_json)

    df = pd.DataFrame(rows)

    excel = to_m6_excel_bytes(df)

    return StreamingResponse(
        BytesIO(excel),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="Declaration_M6.xlsx"'
        },
    )

@app.post("/analyze-acr/start")
async def start_analyze_acr(
    background_tasks: BackgroundTasks,
    video: UploadFile = File(...),
    mapping: UploadFile | None = File(None),
    conductor_type: str = Form("Lundi-Jeudi"),
    intro_type: str = Form("NONE"),
    clean_audio: str = Form("NO"),
):
    job_id = uuid.uuid4().hex
    job_dir = JOBS_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    video_suffix = (
        Path(video.filename or "video.mp4").suffix
        or ".mp4"
    )

    video_path = job_dir / f"video{video_suffix}"

    with video_path.open("wb") as destination:
        while chunk := await video.read(1024 * 1024):
            destination.write(chunk)

    mapping_path = ""

    if mapping is not None:
        mapping_suffix = (
            Path(mapping.filename or "mapping.xlsx").suffix
            or ".xlsx"
        )

        mapping_file = (
            job_dir / f"mapping{mapping_suffix}"
        )

        with mapping_file.open("wb") as destination:
            while chunk := await mapping.read(
                1024 * 1024
            ):
                destination.write(chunk)

        mapping_path = str(mapping_file)

    parameters = {
        "conductor_type": conductor_type,
        "intro_type": intro_type,
        "clean_audio": clean_audio,
    }

    now = utc_now()

    with sqlite3.connect(JOBS_DB) as conn:
        conn.execute(
            """
            INSERT INTO analysis_jobs (
                id,
                status,
                progress,
                current_chunk,
                total_chunks,
                message,
                video_path,
                mapping_path,
                parameters_json,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                job_id,
                "pending",
                0,
                0,
                0,
                "Préparation de l’analyse",
                str(video_path),
                mapping_path,
                json.dumps(
                    parameters,
                    ensure_ascii=False,
                ),
                now,
                now,
            ),
        )

        conn.commit()

    background_tasks.add_task(
        run_acr_job,
        job_id,
    )

    return {
        "job_id": job_id,
        "status": "pending",
    }
@app.get("/jobs/{job_id}")
def read_job(job_id: str):
    job = get_job(job_id)

    if job is None:
        raise HTTPException(
            status_code=404,
            detail="Job introuvable",
        )

    return {
        "id": job["id"],
        "status": job["status"],
        "progress": job["progress"],
        "current_chunk": job["current_chunk"],
        "total_chunks": job["total_chunks"],
        "message": job["message"],
        "result": job["result"],
        "error": job["error"],
        "created_at": job["created_at"],
        "updated_at": job["updated_at"],
    }


@app.post("/jobs/{job_id}/cancel")
def cancel_job(job_id: str):
    if get_job(job_id) is None:
        raise HTTPException(
            status_code=404,
            detail="Job introuvable",
        )

    update_job(
        job_id,
        cancel_requested=1,
        message="Interruption demandée",
    )

    return {"status": "ok"}


@app.post("/jobs/{job_id}/resume")
def resume_job(
    job_id: str,
    background_tasks: BackgroundTasks,
):
    job = get_job(job_id)

    if job is None:
        raise HTTPException(
            status_code=404,
            detail="Job introuvable",
        )

    update_job(
        job_id,
        status="pending",
        cancel_requested=0,
        error=None,
        message="Reprise de l’analyse",
    )

    background_tasks.add_task(
        run_acr_job,
        job_id,
    )

    return {
        "status": "ok",
        "job_id": job_id,
    }
@app.delete("/projects/{project_id}")
def delete_project(project_id: int):
    with sqlite3.connect(PROJECTS_DB) as conn:
        cursor = conn.execute(
            """
            DELETE FROM projects
            WHERE id = ?
            """,
            (project_id,),
        )

        conn.commit()

        if cursor.rowcount == 0:
            raise HTTPException(
                status_code=404,
                detail="Projet introuvable",
            )

    return {
        "ok": True,
        "project_id": project_id,
    }

if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        app,
        host="127.0.0.1",
        port=8000,
        log_level="info",
    )
